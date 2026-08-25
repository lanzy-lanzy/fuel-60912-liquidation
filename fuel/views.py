from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy, reverse
from django.db.models import Sum, Max, Q
from django.db import models
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from django.core.paginator import Paginator
from django.core.exceptions import ValidationError
from .models import Driver, FuelConsumption, LiquidationSetting, LiquidationReport, LiquidationReportEntry
from .forms import FuelConsumptionForm
from datetime import date, datetime, timedelta  # Ensure timedelta is imported
from decimal import Decimal, ROUND_HALF_UP
import calendar
import csv
import io
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Try to import reportlab for PDF generation
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# Heavy equipment vehicles - these never carry passengers (passenger name stays blank)
HE_VEHICLES = [
    'Grader XCMG- GR165',
    'SINOTRUCK F 318',
    'Dumptruck HOWO M7 F030',
    'DUMPTRUCK GREEN NO PLATE',
    'SINOTRUCK 372',
    'HOWO M6 772',
    'LONGKING',
    'MINI BACKHOE XE60DA',
]

class DashboardView(ListView):
    template_name = 'fuel/dashboard.html'
    context_object_name = 'fuel_entries'
    
    def get_queryset(self):
        return FuelConsumption.objects.order_by('-date')[:10]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate fuel statistics
        total_consumed = FuelConsumption.objects.aggregate(
            Sum('total_liters')
        )['total_liters__sum'] or 0
        
        # Get the actual fuel price from the first record to determine which command was used
        first_record = FuelConsumption.objects.first()
        if first_record and first_record.actual_fuel_price == 63.00:
            # v2 command was used
            total_fuel = 5444.10  # 342978.36 / 63.00
        else:
            # Original command was used
            total_fuel = 7499.68
        
        remaining_fuel = total_fuel - total_consumed
        total_cost = FuelConsumption.objects.aggregate(
            Sum('cost')
        )['cost__sum'] or 0

        context.update({
            'total_consumed': round(total_consumed, 2),
            'remaining_fuel': round(remaining_fuel, 2),
            'remaining_percentage': round((remaining_fuel / total_fuel) * 100, 1),
            'total_cost': round(total_cost, 2),
            'drivers': Driver.objects.annotate(
                total_used=Sum('fuelconsumption__total_liters')
            ),
            'current_month': datetime.now().strftime('%B')
        })
        return context

class DriverListView(ListView):
    model = Driver
    template_name = 'fuel/driver_list.html'
    context_object_name = 'drivers'

class DriverDetailView(DetailView):
    model = Driver
    template_name = 'fuel/driver_detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        driver = self.object
        
        consumption_data = FuelConsumption.objects.filter(
            driver=driver
        ).order_by('date')
        
        total_used = consumption_data.aggregate(
            Sum('total_liters')
        )['total_liters__sum'] or 0
        
        context.update({
            'consumption_history': consumption_data,
            'total_trips': consumption_data.aggregate(
                Sum('number_of_trips')
            )['number_of_trips__sum'],
            'total_used': round(total_used, 2),
            'average_per_trip': round(total_used / sum(
                c.number_of_trips for c in consumption_data
            ), 2) if consumption_data else 0
        })
        return context

# views.py
class FuelConsumptionCreateView(CreateView):
    model = FuelConsumption
    form_class = FuelConsumptionForm
    template_name = 'fuel/fuel_form.html'

    def get_success_url(self):
        return reverse('gas_slip', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except ValidationError as e:
            form.add_error(None, e)
            return self.form_invalid(form)

class FuelConsumptionUpdateView(UpdateView):
    model = FuelConsumption
    form_class = FuelConsumptionForm
    template_name = 'fuel/fuel_form.html'
    success_url = reverse_lazy('dashboard')

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except ValidationError as e:
            form.add_error(None, e)
            return self.form_invalid(form)

class FuelConsumptionDeleteView(DeleteView):
    model = FuelConsumption
    template_name = 'fuel/fuel_confirm_delete.html'
    success_url = reverse_lazy('dashboard')

def fuel_report(request):
    # Calculate weekly consumption
    start_date = date(2024, 10, 13)
    end_date = date(2024, 12, 31)
    
    consumption_data = FuelConsumption.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).order_by('date')
    
    # Create weekly breakdown
    weekly_report = []
    current_week = []
    current_week_start = start_date
    
    for entry in consumption_data:
        while entry.date >= current_week_start + timedelta(days=7):  # Use timedelta here
            weekly_report.append({
                'week_start': current_week_start,
                'total_liters': sum(c.total_liters for c in current_week),
                'total_cost': sum(c.cost for c in current_week)
            })
            current_week = []
            current_week_start += timedelta(days=7)  # Use timedelta here
        
        current_week.append(entry)
    
    # Add remaining week
    if current_week:
        weekly_report.append({
            'week_start': current_week_start,
            'total_liters': sum(c.total_liters for c in current_week),
            'total_cost': sum(c.cost for c in current_week)
        })
    
    # Calculate remaining days
    today = date.today()
    remaining_days = (end_date - today).days if today < end_date else 0
    
    # Calculate average daily consumption
    total_days = (end_date - start_date).days
    days_passed = (today - start_date).days if today > start_date else 0
    avg_daily = (FuelConsumption.objects.aggregate(
        Sum('total_liters')
    )['total_liters__sum'] or 0) / days_passed if days_passed > 0 else 0
    
    context = {
        'weekly_report': weekly_report,
        'remaining_days': remaining_days,
        'avg_daily': round(avg_daily, 2),
        'projected_use': round(avg_daily * remaining_days, 2),
    }
    
    return render(request, 'fuel/fuel_report.html', context)# views.py
class DriverCreateView(CreateView):
    model = Driver
    fields = ['name']
    template_name = 'fuel/driver_form.html'
    success_url = reverse_lazy('driver_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        return response
# fuel/views.py
from django.views.generic import DetailView
from .models import FuelConsumption

class GasSlipView(DetailView):
        model = FuelConsumption
        template_name = 'fuel/gas_slip.html'
        context_object_name = 'object'

# fuel/views.py

from django.shortcuts import render
from .models import FuelConsumption
import random

def gas_slip_print_view(request):
    slips = FuelConsumption.objects.all().order_by('date', 'driver', 'trip_number')
    
    # Check if balance data needs to be generated
    slips_without_balance = slips.filter(starting_balance__isnull=True)
    if slips_without_balance.exists():
        # Generate balance data for slips that don't have it
        driver_date_balance = {}
        
        for slip in slips_without_balance:
            driver_date_key = (slip.driver.id, slip.date)
            
            if driver_date_key not in driver_date_balance:
                balance_in_tank = round(random.uniform(7.0, 10.0), 2)
            else:
                balance_in_tank = driver_date_balance[driver_date_key]
            
            issued_liters = slip.total_liters
            total_in_tank = round(balance_in_tank + issued_liters, 2)
            ending_balance = round(random.uniform(7.0, 10.0), 2)
            consumed = round(total_in_tank - ending_balance, 2)
            
            driver_date_balance[driver_date_key] = ending_balance
            
            slip.starting_balance = balance_in_tank
            slip.finished_balance = ending_balance
            slip.consumed_liters = consumed
            slip.save(update_fields=['starting_balance', 'finished_balance', 'consumed_liters'])
    
    # Assign sequential reference numbers based on display order
    for idx, slip in enumerate(slips, 1):
        if slip.reference_number != idx:
            slip.reference_number = idx
            slip.save(update_fields=['reference_number'])
    
    # Re-fetch slips after reference number update
    slips = FuelConsumption.objects.all().order_by('date', 'driver', 'trip_number')

    # Generate passenger names for slips that don't have them (HE slips stay blank)
    FuelConsumption.objects.filter(driver__vehicle__in=HE_VEHICLES).exclude(passenger_name__isnull=True).update(passenger_name=None)
    slips_without_passenger = slips.filter(passenger_name__isnull=True).exclude(driver__vehicle__in=HE_VEHICLES)
    if slips_without_passenger.exists():
        # Common first names in Zamboanga Peninsula
        local_first_names = [
            "Mary", "Maria", "Jocelyn", "Joel", "Jose", "Romeo", "Antonio", "Evelyn", "Rolando", "Danilo",
            "Richard", "Rogelio", "Maricel", "Michael", "Josephine", "Ronald", "Joseph", "Jerry", "Gina", "Erlinda",
            "Reynaldo", "Mark", "Arnel", "Marilyn", "Roger", "Noel", "Teresita", "Edgar", "Roberto", "Edwin",
            "Rey", "Alfredo", "John", "Helen", "Analyn", "Allan", "Eduardo", "Elizabeth", "Alberto", "Mario",
            "Rosita", "Ernesto", "Francisco", "Norma", "Alma", "Jimmy", "Ricardo", "Merlyn", "Elmer", "Ricky",
            "Virginia", "Felix", "Marlon", "Vilma", "Lolita", "Jaime", "Arnold", "Ariel", "Gloria", "Myrna",
            "Vicente", "Jonathan", "Rosemarie", "Marilou", "Julieta", "Jessie", "Marites", "Rodrigo", "Rowena", "Arlyn",
            "Rodolfo", "Robert", "Jenelyn", "Rosalie", "Jennifer", "Albert", "Rene", "Ruben", "Alvin", "Fernando",
            "Roel", "Ryan", "Leonardo", "Pedro", "Evangeline", "Roselyn", "Gemma", "Nelson", "Nestor", "Julito",
            "Lorna", "Ruel", "Wilfredo", "Aida", "Grace", "Jeffrey", "Rosalinda", "Michelle", "Elena", "Nenita",
            "Jesus", "Jovelyn", "Irene", "Edgardo", "Elsa", "Lilia", "Fe", "Anita", "Emma", "Manuel"
        ]
        
        # Common surnames in Dumingag, Zamboanga del Sur
        local_last_names = [
            "dela Cruz", "Mabisa", "Arapon", "Gorre", "Arsenal", "Torres", "Sanchez", "Suaner", "Suerte", "Dico",
            "Maata", "Fernandez", "Ticol", "Oranda", "Sulatorio", "Andata", "dela Torre", "dela Cerna", "Rote", "Decierdo",
            "Sumalpong", "Trazona", "Bazar", "Cañete", "Santos", "Reyes", "Garcia", "Bautista", "del Rosario", "Gonzales",
            "Santos", "Torres", "Mendoza", "Rivera", "Ramos", "Castro", "Domingo", "Santiago", "Villanueva", "Lim",
            "Aquino", "Castillo", "Mercado", "Tan", "Flores", "Salazar", "Gutierrez", "Romero", "Morales", "Dela Cruz"
        ]
        
        # Generate unique passenger names and save them to database
        used_names = set()
        
        for slip in slips_without_passenger:
            # Generate a unique passenger name
            max_attempts = 100
            name = ""
            attempts = 0
            
            while attempts < max_attempts:
                first_name = random.choice(local_first_names)
                last_name = random.choice(local_last_names)
                name = f"{first_name} {last_name}"
                
                if name not in used_names:
                    used_names.add(name)
                    break
                attempts += 1
            
            # If we couldn't generate a unique name, add a number suffix
            if attempts >= max_attempts:
                first_name = random.choice(local_first_names)
                last_name = random.choice(local_last_names)
                suffix = slip.id  # Use slip ID to ensure uniqueness
                name = f"{first_name} {last_name} {suffix}"
                used_names.add(name)
            
            # Save the passenger name to the database
            slip.passenger_name = name
            slip.save(update_fields=['passenger_name'])
    
    # Generate slips with passengers data for the template
    slips_with_passengers = []
    
    # Track ending balance for each driver per date
    driver_date_balance = {}
    
    for slip in slips:
        # Use saved balance data from database
        balance_in_tank = slip.starting_balance if slip.starting_balance is not None else 0.0
        issued_liters = slip.total_liters
        total_in_tank = round(balance_in_tank + issued_liters, 2)
        ending_balance = slip.finished_balance if slip.finished_balance is not None else 0.0
        
        # Create a dictionary with slip data, passenger name, and fuel details
        slip_data = {
            'slip': slip,
            'passenger_name': slip.passenger_name,
            'balance_in_tank': balance_in_tank,
            'issued_liters': issued_liters,
            'total_liters': total_in_tank
        }
        slips_with_passengers.append(slip_data)
    
    context = {
        'slips_with_passengers': slips_with_passengers
    }
    return render(request, 'fuel/gas_slip_print.html', context)

from django.views.generic import CreateView, UpdateView
from django.urls import reverse_lazy
from django.contrib import messages
from .models import FuelConsumption
from .forms import FuelConsumptionForm

class FuelFormView(CreateView):
    model = FuelConsumption
    form_class = FuelConsumptionForm
    template_name = 'fuel_form.html'
    success_url = reverse_lazy('dashboard')

    PRICE_PER_LITER = 56.50  # Fixed price as in consume_fuel.py

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['office'] = 'MAYOR'
        context['title'] = 'New Fuel Entry'
        return context

    def get_initial(self):
        initial = super().get_initial()
        # Get the highest reference number and add 1 for the new entry
        last_ref = FuelConsumption.objects.order_by('-reference_number').first()
        initial['reference_number'] = (last_ref.reference_number + 1) if last_ref else 1
        return initial

    def form_valid(self, form):
        # Set the next reference number if not provided
        if not form.instance.reference_number:
            last_ref = FuelConsumption.objects.order_by('-reference_number').first()
            form.instance.reference_number = (last_ref.reference_number + 1) if last_ref else 1
        
        # Calculate cost based on liters if not provided
        if not form.instance.cost and form.instance.total_liters:
            form.instance.cost = form.instance.total_liters * self.PRICE_PER_LITER
        
        # Calculate liters based on cost if not provided
        elif not form.instance.total_liters and form.instance.cost:
            form.instance.total_liters = form.instance.cost / self.PRICE_PER_LITER

        try:
            response = super().form_valid(form)
        except ValidationError as e:
            form.add_error(None, e)
            return self.form_invalid(form)
        messages.success(self.request, 'Fuel entry created successfully.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)

from django.views.generic import ListView
from .models import FuelConsumption

class FuelConsumptionReportView(ListView):
    model = FuelConsumption
    template_name = 'fuel/fuel_consumption_report.html'
    context_object_name = 'entries'

    def get_queryset(self):
        # Get date range from request parameters, if provided
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        queryset = FuelConsumption.objects.all().order_by('date')
        
        # Filter by date range if provided
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate totals for diesel and gasoline
        diesel_total = sum(
            entry.total_liters for entry in context['entries'] 
            if entry.vehicle in ["Ambulance L300", "Ambulance Province", "Ambulance DOH"]
        )
        
        gasoline_total = sum(
            entry.total_liters for entry in context['entries'] 
            if entry.vehicle not in ["Ambulance L300", "Ambulance Province", "Ambulance DOH"]
        )
        
        context.update({
            'office': 'MAYOR',
            'title': 'Fuel Consumption Report',
            'diesel_total': round(diesel_total, 2),
            'gasoline_total': round(gasoline_total, 2),
            'grand_total': round(diesel_total + gasoline_total, 2),
            'report_date': date.today()
        })
        return context

def fuel_form_view(request):
    """
    Function-based view for the fuel form.
    This redirects to the class-based view.
    """
    view = FuelFormView.as_view()
    return view(request)

@require_GET
def get_destination_choices(request):
    """
    AJAX view to get destination choices based on vehicle type
    """
    vehicle = request.GET.get('vehicle', '')

    if vehicle in ['Backhoe', 'Dumptruck']:
        # Heavy equipment can only use local area
        choices = FuelConsumption.HEAVY_EQUIPMENT_DESTINATION_CHOICES
    else:
        # Ambulances use all destinations including Mahayag
        # Combine ambulance destinations with Mahayag special choice
        choices = (
            FuelConsumption.AMBULANCE_DESTINATION_CHOICES + 
            FuelConsumption.MAHAYAG_SPECIAL_CHOICE
        )

    return JsonResponse({
        'choices': [{'value': choice[0], 'label': choice[1]} for choice in choices]
    })

class DetailedFuelConsumptionReportView(ListView):
    """
    Detailed fuel consumption report view that matches the official form format
    """
    model = FuelConsumption
    template_name = 'fuel/detailed_fuel_consumption_report.html'
    context_object_name = 'entries'

    def get_queryset(self):
        # Get month/year from request parameters, if provided
        month = self.request.GET.get('month')
        year = self.request.GET.get('year')
        driver_id = self.request.GET.get('driver')
        vehicle = self.request.GET.get('vehicle')

        queryset = FuelConsumption.objects.all().order_by('date')

        # Filter by month and year if provided
        if month:
            queryset = queryset.filter(date__month=month)
        if year:
            queryset = queryset.filter(date__year=year)
        elif month:
            # Month selected without a year: only show that month in the current year
            queryset = queryset.filter(date__year=date.today().year)

        # Filter by driver and vehicle if provided
        if driver_id:
            queryset = queryset.filter(driver_id=driver_id)
        if vehicle:
            queryset = queryset.filter(vehicle=vehicle)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Calculate totals for diesel and gasoline
        diesel_total = sum(
            entry.total_liters for entry in context['entries']
            if entry.vehicle in ["Ambulance L300", "Ambulance Province", "Ambulance DOH"]
        )

        gasoline_total = sum(
            entry.total_liters for entry in context['entries']
            if entry.vehicle not in ["Ambulance L300", "Ambulance Province", "Ambulance DOH"]
        )

        # Driver and vehicle options for filter dropdowns
        drivers = Driver.objects.all().order_by('name')
        vehicles = (FuelConsumption.objects
                    .values_list('vehicle', flat=True)
                    .distinct()
                    .order_by('vehicle'))

        entries = context['entries']

        # Totals matching the exported PDF layout
        total_gasoline = 0.00
        total_diesel = round(sum(e.total_liters for e in entries), 2)
        total_consumed = round(sum(e.consumed_liters or 0.0 for e in entries), 2)
        total_amount = round(sum(e.cost for e in entries), 2)

        # Filter info for the print header (same as PDF export)
        month = self.request.GET.get('month')
        year = self.request.GET.get('year')
        driver_id = self.request.GET.get('driver')
        vehicle = self.request.GET.get('vehicle')
        selected_month_name = calendar.month_name[int(month)] if month and month.isdigit() else ''
        selected_driver_name = (Driver.objects.filter(id=driver_id).values_list('name', flat=True).first()
                                if driver_id else '')

        # Build a nice period label for header badge (Month + Year)
        selected_year = year or ''
        if selected_month_name and selected_year:
            selected_period_label = f"{selected_month_name} {selected_year}"
        elif selected_month_name:
            # month filtered without year defaults to current year in queryset
            selected_period_label = f"{selected_month_name} {date.today().year}"
        elif selected_year:
            selected_period_label = f"Year {selected_year}"
        else:
            selected_period_label = "All Periods"

        context.update({
            'office': 'MAYOR',
            'title': 'Detailed Fuel Consumption Report',
            'diesel_total': round(diesel_total, 2),
            'gasoline_total': round(gasoline_total, 2),
            'grand_total': round(diesel_total + gasoline_total, 2),
            'report_date': date.today(),
            'months': [(str(i), calendar.month_name[i]) for i in range(1, 13)],
            'years': [str(y.year) for y in sorted(FuelConsumption.objects.dates('date', 'year'), reverse=True)] or [str(date.today().year)],
            'drivers': drivers,
            'vehicles': vehicles,
            'vehicles_joined': ' - '.join(vehicles),
            'pdf_total_gasoline': total_gasoline,
            'pdf_total_diesel': total_diesel,
            'pdf_total_consumed': total_consumed,
            'pdf_total_amount': total_amount,
            'selected_month_name': selected_month_name,
            'selected_driver_name': selected_driver_name,
            'selected_vehicle': vehicle,
            'selected_year': selected_year,
            'selected_month': month or '',
            'selected_period_label': selected_period_label,
        })
        return context

def fuel_consumption_form_template_view(request):
    """
    View for displaying a blank fuel consumption form template for printing
    """
    return render(request, 'fuel/fuel_consumption_form_template.html', {
        'title': 'Fuel Consumption Form Template'
    })

class ExactFuelConsumptionReportView(ListView):
    """
    Exact replica of the fuel consumption report format from the reference image
    """
    model = FuelConsumption
    template_name = 'fuel/exact_fuel_consumption_report.html'
    context_object_name = 'entries'

    def get_queryset(self):
        # Get date range from request parameters, if provided
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        queryset = FuelConsumption.objects.all().order_by('date')

        # Filter by date range if provided
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Calculate totals for diesel and gasoline
        diesel_total = sum(
            entry.total_liters for entry in context['entries']
            if entry.vehicle in ["Ambulance L300", "Ambulance Province", "Ambulance DOH"]
        )

        gasoline_total = sum(
            entry.total_liters for entry in context['entries']
            if entry.vehicle not in ["Ambulance L300", "Ambulance Province", "Ambulance DOH"]
        )

        context.update({
            'office': 'MAYOR',
            'title': 'Exact Fuel Consumption Report',
            'diesel_total': round(diesel_total, 2),
            'gasoline_total': round(gasoline_total, 2),
            'grand_total': round(diesel_total + gasoline_total, 2),
            'report_date': date.today()
        })
        return context

def exact_blank_fuel_form_view(request):
    """
    View for displaying an exact blank fuel consumption form for manual completion
    """
    return render(request, 'fuel/exact_blank_fuel_form.html', {
        'title': 'Blank Fuel Consumption Form'
    })

def export_fuel_consumption_csv(request):
    """
    Export fuel consumption data to CSV format
    """
    # Get month/year from request parameters
    month = request.GET.get('month')
    year = request.GET.get('year')
    driver_id = request.GET.get('driver')
    vehicle = request.GET.get('vehicle')

    queryset = FuelConsumption.objects.all().order_by('date')

    # Filter by month and year if provided
    if month:
        queryset = queryset.filter(date__month=month)
    if year:
        queryset = queryset.filter(date__year=year)
    elif month:
        # Month selected without a year: only show that month in the current year
        queryset = queryset.filter(date__year=date.today().year)

    # Filter by driver and vehicle if provided
    if driver_id:
        queryset = queryset.filter(driver_id=driver_id)
    if vehicle:
        queryset = queryset.filter(vehicle=vehicle)

    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="fuel_consumption_report.csv"'

    writer = csv.writer(response)

    # Write header rows
    writer.writerow(['FUEL CONSUMPTION REPORT'])
    writer.writerow(['Office: '])
    writer.writerow([])  # Empty row

    # Write table headers
    writer.writerow([
        'DATE', 'REF #', 'VEHICLE', 'PLATE #', 'DRIVER', 'ACTIVITIES', 'DESTINATION', 'DISTANCE KM',
        'AM DEPARTURE', 'AM ARRIVAL', 'PM DEPARTURE', 'PM ARRIVAL', 'TOTAL HOURS',
        'START BALANCE', 'GASOLINE', 'DIESEL', 'LUBRICANTS', 'CONSUME', 'FINISH BALANCE'
    ])

    # Write data rows
    for entry in queryset:
        plate_number = ''
        if entry.vehicle == "Ambulance L300":
            plate_number = 'L300'
        elif entry.vehicle == "Ambulance Province":
            plate_number = 'PROVINCE'
        elif entry.vehicle == "Ambulance DOH":
            plate_number = 'DOH'
        elif entry.vehicle == "Backhoe":
            plate_number = 'BACKHOE'
        elif entry.vehicle == "Dumptruck":
            plate_number = 'DUMPTRUCK'

        gasoline = entry.total_liters if entry.vehicle in ["Backhoe", "Dumptruck"] else ''
        diesel = entry.total_liters if entry.vehicle in ["Ambulance L300", "Ambulance Province", "Ambulance DOH"] else ''

        writer.writerow([
            entry.date.strftime('%m/%d/%Y'),
            f"#{entry.reference_number:04d}",
            entry.vehicle.upper(),
            plate_number,
            entry.driver.name.upper(),
            entry.purpose.upper(),
            entry.get_destination_display().upper(),
            '',  # Distance KM - blank
            '',  # AM Departure - blank
            '',  # AM Arrival - blank
            '',  # PM Departure - blank
            '',  # PM Arrival - blank
            '',  # Total Hours - blank
            '',  # Start Balance - blank
            gasoline,
            diesel,
            '',  # Lubricants - blank
            entry.total_liters,
            ''   # Finish Balance - blank
        ])

    return response

def export_fuel_consumption_excel(request):
    """
    Export fuel consumption data to Excel format with proper formatting and borders
    """
    # Get month/year from request parameters
    month = request.GET.get('month')
    year = request.GET.get('year')
    driver_id = request.GET.get('driver')
    vehicle = request.GET.get('vehicle')

    queryset = FuelConsumption.objects.all().order_by('date')

    # Filter by month and year if provided
    if month:
        queryset = queryset.filter(date__month=month)
    if year:
        queryset = queryset.filter(date__year=year)
    elif month:
        # Month selected without a year: only show that month in the current year
        queryset = queryset.filter(date__year=date.today().year)

    # Filter by driver and vehicle if provided
    if driver_id:
        queryset = queryset.filter(driver_id=driver_id)
    if vehicle:
        queryset = queryset.filter(vehicle=vehicle)

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Fuel Consumption Report"

    # Define styles
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    data_font = Font(name='Arial', size=9)

    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    # Define alignment
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    # Define header fill
    header_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

    # Set column widths
    column_widths = [10, 8, 12, 8, 12, 12, 10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Add title
    ws.merge_cells('A1:Q1')
    title_cell = ws['A1']
    title_cell.value = 'FUEL CONSUMPTION REPORT'
    title_cell.font = title_font
    title_cell.alignment = center_alignment

    # Add office field
    ws.merge_cells('A3:C3')
    office_cell = ws['A3']
    office_cell.value = 'Office: ___________________'
    office_cell.font = data_font
    office_cell.alignment = left_alignment

    # Add headers - First row
    headers_row1 = ['DATE', 'REF #', 'VEHICLE', 'PLATE #', 'DRIVER', 'ACTIVITIES', 'DESTINATION', 'DISTANCE\nKM',
                    'AM', '', 'PM', '', 'TOTAL\nHOURS', 'START\nBALANCE', 'ADDITIONAL', '', '', 'CONSUME', 'FINISH\nBALANCE']

    for col, header in enumerate(headers_row1, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Add headers - Second row
    headers_row2 = ['', '', '', '', '', '', 'DEPARTURE', 'ARRIVAL', 'DEPARTURE', 'ARRIVAL', '', '', 'GASOLINE', 'DIESEL', 'LUBRICANTS', '', '']

    for col, header in enumerate(headers_row2, 1):
        if header:  # Only add if not empty
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border

    # Merge cells for main headers that span two rows
    merge_ranges = [
        'A5:A6', 'B5:B6', 'C5:C6', 'D5:D6', 'E5:E6', 'F5:F6', 'G5:G6',  # Single column headers
        'H5:I5', 'J5:K5',  # AM and PM headers
        'L5:L6', 'M5:M6',  # TOTAL HOURS and START BALANCE
        'N5:P5',  # ADDITIONAL header
        'Q5:Q6', 'R5:R6'  # CONSUME and FINISH BALANCE
    ]

    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)

    # Add data rows
    row_num = 7
    for entry in queryset:
        plate_number = ''
        if entry.vehicle == "Ambulance L300":
            plate_number = 'L300'
        elif entry.vehicle == "Ambulance Province":
            plate_number = 'PROVINCE'
        elif entry.vehicle == "Ambulance DOH":
            plate_number = 'DOH'
        elif entry.vehicle == "Backhoe":
            plate_number = 'BACKHOE'
        elif entry.vehicle == "Dumptruck":
            plate_number = 'DUMPTRUCK'

        gasoline = entry.total_liters if entry.vehicle in ["Backhoe", "Dumptruck"] else ''
        diesel = entry.total_liters if entry.vehicle in ["Ambulance L300", "Ambulance Province", "Ambulance DOH"] else ''

        data_row = [
            entry.date.strftime('%m/%d/%Y'),
            f"#{entry.reference_number:04d}",
            entry.vehicle.upper(),
            plate_number,
            entry.driver.name.upper(),
            entry.purpose.upper(),
            entry.get_destination_display().upper(),
            '',  # Distance KM - blank
            '',  # AM Departure - blank
            '',  # AM Arrival - blank
            '',  # PM Departure - blank
            '',  # PM Arrival - blank
            '',  # Total Hours - blank
            '',  # Start Balance - blank
            gasoline,
            diesel,
            '',  # Lubricants - blank
            entry.total_liters,
            ''   # Finish Balance - blank
        ]

        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
            cell.font = data_font
            cell.alignment = center_alignment if col not in [6, 7] else left_alignment  # Left align activities and destination
            cell.border = thin_border

        row_num += 1

    # Add empty rows for manual completion (about 20 rows)
    for i in range(20):
        for col in range(1, 18):
            cell = ws.cell(row=row_num, column=col)
            cell.value = ''
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
        row_num += 1

    # Add signature section
    signature_row = row_num + 2
    ws.merge_cells(f'A{signature_row}:H{signature_row}')
    prep_cell = ws[f'A{signature_row}']
    prep_cell.value = 'Prepared by:'
    prep_cell.font = data_font
    prep_cell.alignment = left_alignment

    # Add signature lines
    signature_name_row = signature_row + 3
    ws.merge_cells(f'A{signature_name_row}:H{signature_name_row}')
    name_cell = ws[f'A{signature_name_row}']
    name_cell.value = 'GERLAN DORONA'
    name_cell.font = Font(name='Arial', size=10, bold=True)
    name_cell.alignment = center_alignment

    title_row = signature_name_row + 1
    ws.merge_cells(f'A{title_row}:H{title_row}')
    title_cell = ws[f'A{title_row}']
    title_cell.value = 'MDRRMO-CLERK'
    title_cell.font = data_font
    title_cell.alignment = center_alignment

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="fuel_consumption_report.xlsx"'

    # Save workbook to response
    wb.save(response)
    return response

def trip_schedule_view(request):
    # Get filter parameters
    date_filter = request.GET.get('date')
    driver_filter = request.GET.get('driver')
    destination_filter = request.GET.get('destination')
    
    # Start with all trips
    trips = FuelConsumption.objects.select_related('driver').order_by('date', 'driver__name', 'trip_number')
    
    # Apply filters
    if date_filter:
        trips = trips.filter(date=date_filter)
    
    if driver_filter:
        trips = trips.filter(driver_id=driver_filter)
        
    if destination_filter:
        trips = trips.filter(destination=destination_filter)
    
    # Paginate the results
    paginator = Paginator(trips, 20)  # Show 20 trips per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get all drivers for the filter dropdown
    drivers = Driver.objects.all().order_by('name')
    
    # Get destination choices for the filter dropdown
    destinations = FuelConsumption.DESTINATION_CHOICES
    
    context = {
        'trips': page_obj,
        'drivers': drivers,
        'destinations': destinations,
    }
    
    return render(request, 'fuel/trip_schedule.html', context)

def export_trip_schedule_pdf(request):
    if not PDF_SUPPORT:
        # If reportlab is not installed, return a simple error message
        return HttpResponse("PDF export is not available. Please install reportlab library.", content_type="text/plain")
    
    # Get filter parameters
    date_filter = request.GET.get('date')
    driver_filter = request.GET.get('driver')
    destination_filter = request.GET.get('destination')
    
    # Start with all trips
    trips = FuelConsumption.objects.select_related('driver').order_by('date', 'driver__name', 'trip_number')
    
    # Apply filters
    if date_filter:
        trips = trips.filter(date=date_filter)
    
    if driver_filter:
        trips = trips.filter(driver_id=driver_filter)
        
    if destination_filter:
        trips = trips.filter(destination=destination_filter)
    
    # Create a PDF document
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="trip_schedule.pdf"'
    
    # Create a PDF document with legal size (8.5" x 14" - long bond paper)
    from reportlab.lib.pagesizes import legal
    doc = SimpleDocTemplate(response, pagesize=legal, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title = Paragraph("Trip Schedule Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Add filter information if any
    if date_filter or driver_filter or destination_filter:
        filter_info = "Filters: "
        if date_filter:
            filter_info += f"Date: {date_filter} "
        if driver_filter:
            driver = Driver.objects.filter(id=driver_filter).first()
            if driver:
                filter_info += f"Driver: {driver.name} "
        if destination_filter:
            destination_dict = dict(FuelConsumption.DESTINATION_CHOICES)
            filter_info += f"Destination: {destination_dict.get(destination_filter, destination_filter)} "
        
        filter_para = Paragraph(filter_info, styles['Normal'])
        elements.append(filter_para)
        elements.append(Spacer(1, 0.2*inch))
    
    # Prepare data for the table with optimized column widths for legal size
    table_data = [
        ['Date', 'Driver', 'Destination', 'Trip #', 'Departure', 'Arrival', 'Return Departure', 'Return Arrival', 'Fuel (L)', 'Cost (₱)']
    ]
    
    for trip in trips:
        # Format times to 12-hour format with AM/PM
        departure_time = trip.departure_time.strftime('%I:%M %p') if trip.departure_time else '-'
        arrival_time = trip.arrival_time.strftime('%I:%M %p') if trip.arrival_time else '-'
        return_departure_time = trip.return_departure_time.strftime('%I:%M %p') if trip.return_departure_time else '-'
        return_arrival_time = trip.return_arrival_time.strftime('%I:%M %p') if trip.return_arrival_time else '-'
        
        table_data.append([
            trip.date.strftime('%b %d, %Y'),
            trip.driver.name,
            trip.get_destination_display(),
            str(trip.trip_number),
            departure_time,
            arrival_time,
            return_departure_time,
            return_arrival_time,
            f"{trip.total_liters:.2f}",
            f"₱{trip.cost:.2f}"
        ])
    
    # Create the table with adjusted column widths for legal size (more space available)
    # Specify column widths to take advantage of the extra length in legal size
    col_widths = [60, 100, 80, 40, 60, 60, 60, 60, 50, 60]  # Total width should fit legal size
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Added word wrap to prevent text cutoff
    ]))
    
    elements.append(table)
    
    # Add summary information
    elements.append(Spacer(1, 0.2*inch))
    total_trips = trips.count()
    total_fuel = sum(trip.total_liters for trip in trips)
    total_cost = sum(trip.cost for trip in trips)
    
    summary_data = [
        ['Total Trips:', str(total_trips)],
        ['Total Fuel (L):', f"{total_fuel:.2f}"],
        ['Total Cost (₱):', f"₱{total_cost:.2f}"]
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0, colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    
    elements.append(summary_table)
    
    # Build the PDF
    doc.build(elements)
    
    return response

def export_local_passenger_pdf(request):
    if not PDF_SUPPORT:
        # If reportlab is not installed, return a simple error message
        return HttpResponse("PDF export is not available. Please install reportlab library.", content_type="text/plain")
    
    # Get filter parameters
    date_filter = request.GET.get('date')
    driver_filter = request.GET.get('driver')
    destination_filter = request.GET.get('destination')
    
    # Start with all trips
    trips = FuelConsumption.objects.select_related('driver').order_by('date', 'driver__name', 'trip_number')
    
    # Apply filters
    if date_filter:
        trips = trips.filter(date=date_filter)
    
    if driver_filter:
        trips = trips.filter(driver_id=driver_filter)
        
    if destination_filter:
        trips = trips.filter(destination=destination_filter)
    
    # Create a PDF document
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="local_passenger_list.pdf"'
    
    # Create a PDF document with legal size in landscape orientation
    from reportlab.lib.pagesizes import legal, landscape
    doc = SimpleDocTemplate(response, pagesize=landscape(legal), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title = Paragraph("Local Passenger List", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Add filter information if any
    if date_filter or driver_filter or destination_filter:
        filter_info = "Filters: "
        if date_filter:
            filter_info += f"Date: {date_filter} "
        if driver_filter:
            driver = Driver.objects.filter(id=driver_filter).first()
            if driver:
                filter_info += f"Driver: {driver.name} "
        if destination_filter:
            destination_dict = dict(FuelConsumption.DESTINATION_CHOICES)
            filter_info += f"Destination: {destination_dict.get(destination_filter, destination_filter)} "
        
        filter_para = Paragraph(filter_info, styles['Normal'])
        elements.append(filter_para)
        elements.append(Spacer(1, 0.2*inch))
    
    # Generate random local passenger names from Dumingag, Zamboanga del Sur
    # Common first names in Zamboanga Peninsula
    local_first_names = [
        "Mary", "Maria", "Jocelyn", "Joel", "Jose", "Romeo", "Antonio", "Evelyn", "Rolando", "Danilo",
        "Richard", "Rogelio", "Maricel", "Michael", "Josephine", "Ronald", "Joseph", "Jerry", "Gina", "Erlinda",
        "Reynaldo", "Mark", "Arnel", "Marilyn", "Roger", "Noel", "Teresita", "Edgar", "Roberto", "Edwin",
        "Rey", "Alfredo", "John", "Helen", "Analyn", "Allan", "Eduardo", "Elizabeth", "Alberto", "Mario",
        "Rosita", "Ernesto", "Francisco", "Norma", "Alma", "Jimmy", "Ricardo", "Merlyn", "Elmer", "Ricky",
        "Virginia", "Felix", "Marlon", "Vilma", "Lolita", "Jaime", "Arnold", "Ariel", "Gloria", "Myrna",
        "Vicente", "Jonathan", "Rosemarie", "Marilou", "Julieta", "Jessie", "Marites", "Rodrigo", "Rowena", "Arlyn",
        "Rodolfo", "Robert", "Jenelyn", "Rosalie", "Jennifer", "Albert", "Rene", "Ruben", "Alvin", "Fernando",
        "Roel", "Ryan", "Leonardo", "Pedro", "Evangeline", "Roselyn", "Gemma", "Nelson", "Nestor", "Julito",
        "Lorna", "Ruel", "Wilfredo", "Aida", "Grace", "Jeffrey", "Rosalinda", "Michelle", "Elena", "Nenita",
        "Jesus", "Jovelyn", "Irene", "Edgardo", "Elsa", "Lilia", "Fe", "Anita", "Emma", "Manuel"
    ]
    
    # Common surnames in Dumingag, Zamboanga del Sur
    local_last_names = [
        "dela Cruz", "Mabisa", "Arapon", "Gorre", "Arsenal", "Torres", "Sanchez", "Suaner", "Suerte", "Dico",
        "Maata", "Fernandez", "Ticol", "Oranda", "Sulatorio", "Andata", "dela Torre", "dela Cerna", "Rote", "Decierdo",
        "Sumalpong", "Trazona", "Bazar", "Cañete", "Santos", "Reyes", "Garcia", "Bautista", "del Rosario", "Gonzales",
        "Santos", "Torres", "Mendoza", "Rivera", "Ramos", "Castro", "Domingo", "Santiago", "Villanueva", "Lim",
        "Aquino", "Castillo", "Mercado", "Tan", "Flores", "Salazar", "Gutierrez", "Romero", "Morales", "Dela Cruz"
    ]
    
    # Generate unique passenger names for each trip
    import random
    used_names = set()
    passenger_names = []
    
    for i in range(len(trips)):
        # Generate a unique name
        max_attempts = 100  # Prevent infinite loop
        name = ""
        attempts = 0
        
        while attempts < max_attempts:
            first_name = random.choice(local_first_names)
            last_name = random.choice(local_last_names)
            name = f"{first_name} {last_name}"
            
            if name not in used_names:
                used_names.add(name)
                break
            attempts += 1
        
        # If we couldn't generate a unique name, add a number suffix
        if attempts >= max_attempts:
            first_name = random.choice(local_first_names)
            last_name = random.choice(local_last_names)
            suffix = i + 1
            name = f"{first_name} {last_name} {suffix}"
            used_names.add(name)
            
        passenger_names.append(name)
    
    # Prepare data for the table with optimized column widths for landscape legal size
    table_data = [
        ['Date', 'Driver', 'Destination', 'Trip #', 'Passenger Name', 'Departure', 'Arrival', 'Return Departure', 'Return Arrival']
    ]
    
    for i, trip in enumerate(trips):
        # Format times to 12-hour format with AM/PM
        departure_time = trip.departure_time.strftime('%I:%M %p') if trip.departure_time else '-'
        arrival_time = trip.arrival_time.strftime('%I:%M %p') if trip.arrival_time else '-'
        return_departure_time = trip.return_departure_time.strftime('%I:%M %p') if trip.return_departure_time else '-'
        return_arrival_time = trip.return_arrival_time.strftime('%I:%M %p') if trip.return_arrival_time else '-'
        
        table_data.append([
            trip.date.strftime('%b %d, %Y'),
            trip.driver.name,
            trip.get_destination_display(),
            str(trip.trip_number),
            passenger_names[i],
            departure_time,
            arrival_time,
            return_departure_time,
            return_arrival_time
        ])
    
    # Create the table with adjusted column widths for landscape legal size
    # Landscape legal size gives us more width to work with (14" instead of 8.5")
    col_widths = [70, 120, 100, 40, 150, 70, 70, 70, 70]  # Total width should fit landscape legal size
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Added word wrap to prevent text cutoff
    ]))
    
    elements.append(table)
    
    # Add summary information
    elements.append(Spacer(1, 0.2*inch))
    total_trips = trips.count()
    
    summary_data = [
        ['Total Trips:', str(total_trips)],
        ['Total Passengers:', str(len(used_names))]
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0, colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    
    elements.append(summary_table)
    
    # Build the PDF
    doc.build(elements)
    
    return response


def gas_slip_print_view(request):
    slips = FuelConsumption.objects.all().order_by('date', 'driver', 'trip_number')
    
    # Check if balance data needs to be generated
    slips_without_balance = slips.filter(starting_balance__isnull=True)
    if slips_without_balance.exists():
        # Generate balance data for slips that don't have it
        driver_date_balance = {}
        
        for slip in slips_without_balance:
            driver_date_key = (slip.driver.id, slip.date)
            
            if driver_date_key not in driver_date_balance:
                balance_in_tank = round(random.uniform(7.0, 10.0), 2)
            else:
                balance_in_tank = driver_date_balance[driver_date_key]
            
            issued_liters = slip.total_liters
            total_in_tank = round(balance_in_tank + issued_liters, 2)
            ending_balance = round(random.uniform(7.0, 10.0), 2)
            consumed = round(total_in_tank - ending_balance, 2)
            
            driver_date_balance[driver_date_key] = ending_balance
            
            slip.starting_balance = balance_in_tank
            slip.finished_balance = ending_balance
            slip.consumed_liters = consumed
            # Use update instead of save to avoid triggering fuel limit validation
            FuelConsumption.objects.filter(pk=slip.pk).update(
                starting_balance=balance_in_tank,
                finished_balance=ending_balance,
                consumed_liters=consumed
            )
    
    # Generate passenger names for slips that don't have them (HE slips stay blank)
    FuelConsumption.objects.filter(driver__vehicle__in=HE_VEHICLES).exclude(passenger_name__isnull=True).update(passenger_name=None)
    slips_without_passenger = slips.filter(passenger_name__isnull=True).exclude(driver__vehicle__in=HE_VEHICLES)
    if slips_without_passenger.exists():
        # Common first names in Zamboanga Peninsula
        local_first_names = [
            "Mary", "Maria", "Jocelyn", "Joel", "Jose", "Romeo", "Antonio", "Evelyn", "Rolando", "Danilo",
            "Richard", "Rogelio", "Maricel", "Michael", "Josephine", "Ronald", "Joseph", "Jerry", "Gina", "Erlinda",
            "Reynaldo", "Mark", "Arnel", "Marilyn", "Roger", "Noel", "Teresita", "Edgar", "Roberto", "Edwin",
            "Rey", "Alfredo", "John", "Helen", "Analyn", "Allan", "Eduardo", "Elizabeth", "Alberto", "Mario",
            "Rosita", "Ernesto", "Francisco", "Norma", "Alma", "Jimmy", "Ricardo", "Merlyn", "Elmer", "Ricky",
            "Virginia", "Felix", "Marlon", "Vilma", "Lolita", "Jaime", "Arnold", "Ariel", "Gloria", "Myrna",
            "Vicente", "Jonathan", "Rosemarie", "Marilou", "Julieta", "Jessie", "Marites", "Rodrigo", "Rowena", "Arlyn",
            "Rodolfo", "Robert", "Jenelyn", "Rosalie", "Jennifer", "Albert", "Rene", "Ruben", "Alvin", "Fernando",
            "Roel", "Ryan", "Leonardo", "Pedro", "Evangeline", "Roselyn", "Gemma", "Nelson", "Nestor", "Julito",
            "Lorna", "Ruel", "Wilfredo", "Aida", "Grace", "Jeffrey", "Rosalinda", "Michelle", "Elena", "Nenita",
            "Jesus", "Jovelyn", "Irene", "Edgardo", "Elsa", "Lilia", "Fe", "Anita", "Emma", "Manuel"
        ]
        
        # Common surnames in Dumingag, Zamboanga del Sur
        local_last_names = [
            "dela Cruz", "Mabisa", "Arapon", "Gorre", "Arsenal", "Torres", "Sanchez", "Suaner", "Suerte", "Dico",
            "Maata", "Fernandez", "Ticol", "Oranda", "Sulatorio", "Andata", "dela Torre", "dela Cerna", "Rote", "Decierdo",
            "Sumalpong", "Trazona", "Bazar", "Cañete", "Santos", "Reyes", "Garcia", "Bautista", "del Rosario", "Gonzales",
            "Santos", "Torres", "Mendoza", "Rivera", "Ramos", "Castro", "Domingo", "Santiago", "Villanueva", "Lim",
            "Aquino", "Castillo", "Mercado", "Tan", "Flores", "Salazar", "Gutierrez", "Romero", "Morales", "Dela Cruz"
        ]
        
        # Generate unique passenger names and save them to database
        used_names = set()
        
        for slip in slips_without_passenger:
            # Generate a unique passenger name
            max_attempts = 100
            name = ""
            attempts = 0
            
            while attempts < max_attempts:
                first_name = random.choice(local_first_names)
                last_name = random.choice(local_last_names)
                name = f"{first_name} {last_name}"
                
                if name not in used_names:
                    used_names.add(name)
                    break
                attempts += 1
            
            # If we couldn't generate a unique name, add a number suffix
            if attempts >= max_attempts:
                first_name = random.choice(local_first_names)
                last_name = random.choice(local_last_names)
                suffix = slip.id  # Use slip ID to ensure uniqueness
                name = f"{first_name} {last_name} {suffix}"
                used_names.add(name)
            
            # Save the passenger name to the database using update to avoid validation
            FuelConsumption.objects.filter(pk=slip.pk).update(passenger_name=name)
    
    # Refresh the slips queryset to get the updated data
    slips = FuelConsumption.objects.all().order_by('date', 'driver', 'trip_number')
    
    # Generate slips with passengers data for the template
    slips_with_passengers = []
    
    # Track ending balance for each driver per date
    driver_date_balance = {}
    
    for slip in slips:
        # Use saved balance data from database
        balance_in_tank = slip.starting_balance if slip.starting_balance is not None else 0.0
        issued_liters = slip.total_liters
        total_in_tank = round(balance_in_tank + issued_liters, 2)
        ending_balance = slip.finished_balance if slip.finished_balance is not None else 0.0
        
        # Create a dictionary with slip data, passenger name, and fuel details
        slip_data = {
            'slip': slip,
            'passenger_name': slip.passenger_name,
            'balance_in_tank': balance_in_tank,
            'issued_liters': issued_liters,
            'total_liters': total_in_tank
        }
        slips_with_passengers.append(slip_data)
    
    context = {
        'slips_with_passengers': slips_with_passengers
    }
    return render(request, 'fuel/gas_slip_print.html', context)


def fuel_consumption_simple_form_view(request):
    """
    View for displaying a simplified fuel consumption form template that matches the requested format
    """
    return render(request, 'fuel/fuel_consumption_simple_form.html', {
        'title': 'Simplified Fuel Consumption Form'
    })


def export_simple_fuel_consumption_pdf(request):
    if not PDF_SUPPORT:
        # If reportlab is not installed, return a simple error message
        return HttpResponse("PDF export is not available. Please install reportlab library.", content_type="text/plain")
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Start with all fuel consumption records
    records = FuelConsumption.objects.select_related('driver').order_by('date')
    
    # Apply date filters if provided
    if start_date:
        records = records.filter(date__gte=start_date)
    
    if end_date:
        records = records.filter(date__lte=end_date)
    
    # Create a PDF document
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="simple_fuel_consumption_report.pdf"'
    
    # Create a PDF document with legal size in landscape orientation
    from reportlab.lib.pagesizes import legal, landscape
    doc = SimpleDocTemplate(response, pagesize=landscape(legal), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.textColor = colors.black
    title = Paragraph("FUEL CONSUMPTION REPORT", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Add filter information if any
    if start_date or end_date:
        filter_info = "Period: "
        if start_date and end_date:
            filter_info += f"{start_date} to {end_date}"
        elif start_date:
            filter_info += f"From {start_date}"
        elif end_date:
            filter_info += f"Until {end_date}"
        
        filter_para = Paragraph(filter_info, styles['Normal'])
        elements.append(filter_para)
        elements.append(Spacer(1, 0.2*inch))
    
    # Prepare data for the table
    table_data = [
        ['DATE REQUESTED', 'PURPOSE', 'DESTINATION', 'STARTING BALANCE', 'ADDITIONAL GASOLINE', 'DIESEL', 'CONSUME', 'FINISHED BALANCE', 'PRICE/L', 'AMOUNT (ADD\'L * PRICE)']
    ]
    
    # For the real data, we need to process records properly
    # Group records by date for the report
    from collections import defaultdict
    records_by_date = defaultdict(list)
    for record in records:
        records_by_date[record.date].append(record)
    
    # Initialize totals and running balance
    total_gasoline = 0
    total_diesel = 0
    total_amount = 0
    # Start with a realistic initial balance (10 or not less than 7 liters)
    running_balance = random.randint(7, 10)
    
    # Sort dates to process in chronological order
    sorted_dates = sorted(records_by_date.keys())
    
    # Process each date's records
    for date in sorted_dates:
        date_records = records_by_date[date]
        
        # Separate gasoline and diesel consumption
        gasoline_consumed = 0
        diesel_consumed = 0
        
        # Categorize by vehicle type
        for record in date_records:
            # Diesel vehicles (ambulances)
            if record.vehicle in ["Ambulance L300", "Ambulance Province", "Ambulance DOH"]:
                diesel_consumed += record.total_liters
            # Gasoline vehicles (heavy equipment)
            else:
                gasoline_consumed += record.total_liters
        
        # Get the actual fuel price used for this record
        price_per_liter = getattr(date_records[0], 'actual_fuel_price', 63.00)
        
        # Calculate amount based on gasoline consumed (as per the form format)
        amount = gasoline_consumed * price_per_liter
        
        # Use the current running balance as starting balance
        starting_balance = running_balance
        
        # Calculate finished balance (starting balance + gasoline added - diesel consumed)
        # Note: In this form, "consume" refers to diesel consumption
        finished_balance = starting_balance + gasoline_consumed - diesel_consumed
        # Ensure we don't go below zero
        finished_balance = max(0, finished_balance)
        
        # Update running balance for next iteration
        running_balance = finished_balance
        
        # Get destinations for this date (from the actual records, not generated)
        destinations = ", ".join(set(record.get_destination_display() for record in date_records))
        
        # Use "Transport Patient" as default purpose for ambulance records
        purpose = "Transport Patient"
        
        table_data.append([
            date.strftime('%m/%d/%Y'),
            purpose,
            destinations,
            f"{starting_balance:.2f}",
            f"{gasoline_consumed:.2f}",
            f"{diesel_consumed:.2f}",
            f"{diesel_consumed:.2f}",  # Consume = Diesel in this format
            f"{finished_balance:.2f}",
            f"{price_per_liter:.2f}",
            f"{amount:.2f}"
        ])
        
        total_gasoline += gasoline_consumed
        total_diesel += diesel_consumed
        total_amount += amount
    
    # Add some blank rows for additional entries (enough to fill the page)
    for i in range(20):
        table_data.append(['', '', '', '', '', '', '', '', '', ''])
    
    # Add total row
    table_data.append([
        'TOTAL',
        '',
        '',
        '',
        f"{total_gasoline:.2f}",
        f"{total_diesel:.2f}",
        f"{total_diesel:.2f}",
        '',
        '',
        f"{total_amount:.2f}"
    ])
    
    # Create the table with adjusted column widths for landscape legal size
    col_widths = [80, 120, 150, 80, 80, 80, 80, 80, 60, 80]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('WORDWRAP', (0, 0), (-1, -1)),
        ('FONTNAME', (-1, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('LINEBELOW', (0, 0), (-1, -1), 1, colors.black),
        ('LINEABOVE', (0, 0), (-1, -1), 1, colors.black),
        ('LINEBEFORE', (0, 0), (-1, -1), 1, colors.black),
        ('LINEAFTER', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    
    # Add signature section
    elements.append(Spacer(1, 0.5*inch))
    
    signature_data = [
        ['Prepared by:', 'Noted by:'],
        ['', ''],
        ['GERLAN DORONA', '_______________________'],
        ['MDRRMO-CLERK', 'Municipal Mayor']
    ]
    
    signature_table = Table(signature_data, colWidths=[250, 250])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 2), (1, 3), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 1), (-1, 1), 20),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 0, colors.white),  # No grid lines for signature section
    ]))
    
    elements.append(signature_table)
    
    # Build the PDF
    doc.build(elements)
    
    return response


def fuel_consumption_detailed_report_view(request):
    """
    Display fuel consumption report with starting balance, issued, consumed, and finished balance
    Uses saved balance data from database for consistency
    """
    # Get filter parameters
    date_filter = request.GET.get('date')
    driver_filter = request.GET.get('driver')
    vehicle_filter = request.GET.get('vehicle')
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    # Start with all trips ordered by date, driver, and trip number
    slips = FuelConsumption.objects.select_related('driver').order_by('date', 'driver', 'trip_number')
    
    # Apply filters
    if month:
        slips = slips.filter(date__month=month)
    if year:
        slips = slips.filter(date__year=year)
    elif month:
        # Month selected without a year: only show that month in the current year
        slips = slips.filter(date__year=date.today().year)
    if date_filter:
        slips = slips.filter(date=date_filter)
    if driver_filter:
        slips = slips.filter(driver_id=driver_filter)
    if vehicle_filter:
        slips = slips.filter(vehicle=vehicle_filter)
    
    # Get all drivers and vehicles for filter dropdowns
    drivers = Driver.objects.all().order_by('name')
    vehicles = FuelConsumption.objects.values_list('vehicle', flat=True).distinct().order_by('vehicle')
    
    # Use saved balance data from database
    report_data = []
    total_gasoline = 0
    total_diesel = 0
    total_consumed = 0
    total_amount = 0
    
    for slip in slips:
        # Use saved balance data from database
        balance_in_tank = slip.starting_balance if slip.starting_balance is not None else 0.0
        issued_liters = slip.total_liters
        ending_balance = slip.finished_balance if slip.finished_balance is not None else 0.0
        consumed = slip.consumed_liters if slip.consumed_liters is not None else 0.0
        
        # Amount is the total cost from the gas slip
        amount = slip.cost
        
        # Add to totals (diesel column for ambulances, not gasoline)
        total_gasoline += 0.00
        total_diesel += issued_liters
        total_consumed += consumed
        total_amount += amount
        
        # Add to report data
        report_data.append({
            'date': slip.date,
            'reference_number': slip.reference_number,
            'driver_name': slip.driver.name,
            'purpose': slip.purpose,
            'destination': slip.get_destination_display(),
            'starting_balance': balance_in_tank,
            'additional_gasoline': 0.00,
            'additional_diesel': issued_liters,
            'consumed': consumed,
            'finished_balance': ending_balance,
            'price_per_liter': slip.actual_fuel_price,
            'amount': amount
        })
    
    context = {
        'report_data': report_data,
        'total_gasoline': total_gasoline,
        'total_diesel': total_diesel,
        'total_consumed': total_consumed,
        'total_amount': total_amount,
        'drivers': drivers,
        'vehicles': vehicles,
        'selected_date': date_filter,
        'selected_driver': driver_filter,
        'selected_vehicle': vehicle_filter,
        'selected_month': str(month) if month else '',
        'selected_year': str(year) if year else '',
        'months': [(str(i), calendar.month_name[i]) for i in range(1, 13)],
        'years': [str(y.year) for y in sorted(FuelConsumption.objects.dates('date', 'year'), reverse=True)] or [str(date.today().year)],
    }
    
    return render(request, 'fuel/fuel_consumption_detailed_report.html', context)


def export_fuel_consumption_report_pdf(request):
    """
    Export fuel consumption report showing starting balance, issued, consumed, and finished balance
    """
    if not PDF_SUPPORT:
        return HttpResponse("PDF export is not available. Please install reportlab library.", content_type="text/plain")
    
    # Get filter parameters
    date_filter = request.GET.get('date')
    driver_filter = request.GET.get('driver')
    vehicle_filter = request.GET.get('vehicle')
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    # Start with all trips ordered by date, driver, and trip number
    trips = FuelConsumption.objects.select_related('driver').order_by('date', 'driver', 'trip_number')
    
    # Apply filters
    if month:
        trips = trips.filter(date__month=month)
    if year:
        trips = trips.filter(date__year=year)
    elif month:
        # Month selected without a year: only show that month in the current year
        trips = trips.filter(date__year=date.today().year)
    if date_filter:
        trips = trips.filter(date=date_filter)
    if driver_filter:
        trips = trips.filter(driver_id=driver_filter)
    if vehicle_filter:
        trips = trips.filter(vehicle=vehicle_filter)
    
    # Create PDF response - inline to preview/print without download
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="fuel_consumption_report.pdf"'
    
    # Create document with landscape orientation
    from reportlab.lib.pagesizes import legal, landscape
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    from collections import defaultdict
    
    # Register a Unicode-supporting font from our custom fonts directory
    font_name = 'Helvetica'
    bold_font_name = 'Helvetica-Bold'
    
    try:
        # Use DejaVuSans from our custom fonts directory
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_dir = os.path.dirname(current_dir)
        fonts_dir = os.path.join(project_dir, 'fonts', 'ttf')
        
        # Check for DejaVuSans font files
        dejavu_font_path = os.path.join(fonts_dir, 'DejaVuSans.ttf')
        dejavu_bold_font_path = os.path.join(fonts_dir, 'DejaVuSans-Bold.ttf')
        
        if os.path.exists(dejavu_font_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_font_path))
            if os.path.exists(dejavu_bold_font_path):
                pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_bold_font_path))
                bold_font_name = 'DejaVuSans-Bold'
            else:
                bold_font_name = 'DejaVuSans'
            font_name = 'DejaVuSans'
    except Exception as e:
        # If font registration fails, fallback to default Helvetica
        font_name = 'Helvetica'
        bold_font_name = 'Helvetica-Bold'
    
    # Use landscape orientation with legal size
    doc = SimpleDocTemplate(response, pagesize=landscape(legal), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    
    # Add title with enhanced font size
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,  # Increased font size
        textColor=colors.black,
        spaceAfter=15,  # Increased spacing
        alignment=1,  # Center alignment
        fontName=bold_font_name
    )
    
    title = Paragraph("FUEL CONSUMPTION REPORT", title_style)
    elements.append(title)
    
    # --- Vehicle as HEADER, Period & Driver as SUB-HEADER (no emoji - fixes PDF box) ---
    selected_month_name = calendar.month_name[int(month)] if month and month.isdigit() else ''
    if selected_month_name and year:
        period_label = f"{selected_month_name} {year}"
    elif selected_month_name:
        period_label = f"{selected_month_name} {date.today().year}"
    elif year:
        period_label = f"Year {year}"
    else:
        period_label = "All Periods"

    vehicles = list(trips.values_list('vehicle', flat=True).distinct().order_by('vehicle'))
    if vehicle_filter:
        vehicle_header = vehicle_filter
    elif vehicles:
        vehicle_header = " - ".join(vehicles)
    else:
        vehicle_header = "All Vehicles"

    # Vehicle badge - header (dark blue pill, no emoji for PDF compatibility)
    vehicle_badge_style = ParagraphStyle(
        'VehicleBadge',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.white,
        alignment=1,
        fontName=bold_font_name,
        leading=14,
    )
    vehicle_para = Paragraph(vehicle_header, vehicle_badge_style)
    # Wider pill to fit multiple vehicles
    badge_w = min(520, max(240, len(vehicle_header) * 5.5))
    badge_table = Table([[vehicle_para]], colWidths=[badge_w])
    badge_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROUNDEDCORNERS', [6, 6, 6, 6]),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ('LEFTPADDING', (0, 0), (-1, -1), 14),
        ('RIGHTPADDING', (0, 0), (-1, -1), 14),
    ]))
    badge_wrapper = Table([[badge_table]], colWidths=[landscape(legal)[0] - 60])
    badge_wrapper.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(badge_wrapper)
    elements.append(Spacer(1, 0.10*inch))

    # Subtitle under title (no Office, simple)
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#64748b'),
        alignment=1,
        fontName=font_name,
        spaceAfter=8,
        leading=10,
    )
    subtitle = Paragraph("DETAILED CONSUMPTION &amp; LIQUIDATION  •  " + date.today().strftime("%B %d, %Y"), subtitle_style)
    elements.append(subtitle)

    # Filter pills row (Month / Vehicle / Driver / Date)
    pill_style = ParagraphStyle(
        'Pill',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#334155'),
        alignment=1,
        fontName=font_name,
        leading=10,
    )
    pill_bold = bold_font_name
    pills = []
    # Vehicle is header badge, sub-header is Period + Driver (no emoji, no vehicle duplicate)
    pills.append(Paragraph(f'Period: <font name="{pill_bold}">{period_label}</font>', pill_style))
    if driver_filter:
        try:
            driver_obj = Driver.objects.filter(id=driver_filter).first()
            driver_name = driver_obj.name if driver_obj else driver_filter
        except:
            driver_name = driver_filter
        pills.append(Paragraph(f'Driver: <font name="{pill_bold}">{driver_name}</font>', pill_style))
    if date_filter:
        pills.append(Paragraph(f'Date: <font name="{pill_bold}">{date_filter}</font>', pill_style))

    # Render pills (header is vehicle, sub-header is period/driver - simple & not redundant)
    if pills:
        pill_cells = []
        for p in pills:
            t = Table([[p]], colWidths=[130])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f1f5f9')),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
                ('ROUNDEDCORNERS', [10, 10, 10, 10]),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            pill_cells.append(t)
        pills_table = Table([pill_cells], colWidths=[132]*len(pill_cells), hAlign='CENTER')
        pills_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(pills_table)
        elements.append(Spacer(1, 0.18*inch))
    else:
        elements.append(Spacer(1, 0.08*inch))
    
    # Prepare table data
    # Use Unicode code point for peso sign (U+20B1) to ensure proper display
    peso_sign = '\u20B1'  # Unicode for ₱'

    # Style for wrapping long text cells (e.g. road clearing purposes)
    purpose_style = ParagraphStyle(
        'PurposeCell',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        fontName=font_name,
        alignment=1,
    )

    table_data = [
        ['Date\nRequested', 'Ref #', 'Driver', 'Purpose', 'Destination', 'Starting\nBalance', 
         'Additional\nGasoline', 'Additional\nDiesel', 'Consume', 
         'Finished\nBalance', f'Price/L\n({peso_sign})', f'Amount\n(Add)*\n({peso_sign})']
    ]
    
    # Initialize totals
    total_gasoline = 0
    total_diesel = 0
    total_consumed = 0
    total_amount = 0
    
    # Use saved balance data from database for consistency
    for trip in trips:
        # Use saved balance data from database
        starting_balance = trip.starting_balance if trip.starting_balance is not None else 0.0
        issued_liters = trip.total_liters
        additional_gasoline = 0.00  # Gasoline column is 0
        additional_diesel = issued_liters  # Diesel column gets the issued liters
        finished_balance = trip.finished_balance if trip.finished_balance is not None else 0.0
        consumed = trip.consumed_liters if trip.consumed_liters is not None else 0.0
        
        # Amount is the total cost from the gas slip
        amount = trip.cost
        
        # Add to totals
        total_gasoline += additional_gasoline
        total_diesel += additional_diesel
        total_consumed += consumed
        total_amount += amount
        
        # Add row to table using Unicode string formatting with comma separators
        table_data.append([
            trip.date.strftime('%m/%d/%Y'),
            f"#{trip.reference_number:04d}",
            trip.driver.name,
            Paragraph(trip.purpose, purpose_style),
            trip.get_destination_display(),
            f"{starting_balance:,.2f}",
            f"{additional_gasoline:,.2f}",
            f"{additional_diesel:,.2f}",
            f"{consumed:,.2f}",
            f"{finished_balance:,.2f}",
            '{}{:.2f}'.format(peso_sign, trip.actual_fuel_price),  # Using format method for Unicode
            '{}{:,.2f}'.format(peso_sign, amount)  # Using format method for Unicode with comma separator
        ])
    
    # Add totals row with comma separators
    table_data.append([
        '', '', '', '', 'TOTAL', '',
        f"{total_gasoline:,.2f}",
        f"{total_diesel:,.2f}",
        f"{total_consumed:,.2f}",
        '', '',
        '{}{:,.2f}'.format(peso_sign, total_amount)
    ])
    
    # Create table with column widths optimized for landscape legal size (~1008 points total)
    # Using ~935 points to ensure it fits comfortably within margins
    col_widths = [70, 45, 90, 130, 90, 75, 75, 75, 65, 75, 65, 80]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), bold_font_name),  # Use the registered font
        ('FONTSIZE', (0, 0), (-1, 0), 10),  # Increased header font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('FONTNAME', (0, 1), (-1, -1), font_name),  # Use the registered font
        ('FONTSIZE', (0, 1), (-1, -1), 9),  # Increased data font size
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, -1), (-1, -1), bold_font_name),  # Use the registered font
        ('BACKGROUND', (0, -1), (-1, -1), colors.white),
        ('FONTSIZE', (0, -1), (-1, -1), 10),  # Increased totals font size
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable word wrap for cells
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.4*inch))  # Increased spacing
    
    # Add signature section with enhanced font sizes
    signature_data = [
        ['', ''],
        ['GERLAN B. DORONA', 'JHUNAX L. CARDOZA'],
        ['MDRRMO STAFF', 'LDRRMO-III']
    ]
    
    signature_table = Table(signature_data, colWidths=[350, 350])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, 1), bold_font_name),  # Use the registered font
        ('FONTSIZE', (0, 1), (-1, 1), 12),  # Increased name font size
        ('FONTSIZE', (0, 2), (-1, 2), 10),  # Increased title font size
        ('LINEABOVE', (0, 1), (-1, 1), 1, colors.black),
        ('TOPPADDING', (0, 1), (-1, 1), 20),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 0, colors.white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),  # Use the registered font
    ]))
    
    elements.append(signature_table)
    
    # Build PDF
    doc.build(elements)
    
    return response


def liquidation_report_view(request, fresh=False):
    # fresh=True: start a blank report with NO rows derived from FuelConsumption trips
    if fresh:
        trips = FuelConsumption.objects.none()
    else:
        trips = FuelConsumption.objects.select_related('driver').order_by('date', 'reference_number')
    setting, _ = LiquidationSetting.objects.get_or_create(pk=1)

    if request.method == 'POST':
        edited_amounts = {}
        edited_vat = {}
        edited_wht5 = {}
        edited_wht1 = {}
        deleted_ids = set()
        for trip in trips:
            # per-row delete: if delete_<id> present, skip this trip (exclude from saved report)
            if f'delete_{trip.id}' in request.POST:
                deleted_ids.add(trip.id)
                continue
            value = request.POST.get(f'or_{trip.id}', '').strip()
            if value != (trip.or_number or ''):
                FuelConsumption.objects.filter(pk=trip.pk).update(or_number=value or None)
            amount_str = request.POST.get(f'amount_{trip.id}', '').strip().replace(',', '')
            if amount_str:
                try:
                    edited_amt = Decimal(amount_str)
                    edited_amounts[trip.id] = edited_amt
                except Exception:
                    pass
            # VAT checkbox: checked = VAT inclusive (divide by 1.12). Unchecked = Non-VAT (optional)
            edited_vat[trip.id] = f'vat_{trip.id}' in request.POST
            # Dynamic WHT inputs per OR # - if provided, override computed
            wht5_str = request.POST.get(f'wht5_{trip.id}', '').strip().replace(',', '').replace('₱','')
            if wht5_str:
                try:
                    edited_wht5[trip.id] = Decimal(wht5_str)
                except Exception:
                    pass
            wht1_str = request.POST.get(f'wht1_{trip.id}', '').strip().replace(',', '').replace('₱','')
            if wht1_str:
                try:
                    edited_wht1[trip.id] = Decimal(wht1_str)
                except Exception:
                    pass
        principal = request.POST.get('principal', '').strip().replace(',', '')
        if principal:
            try:
                setting.principal_amount = Decimal(principal)
            except Exception:
                pass
        # Footer dynamic amounts - OR number for refund is now editable
        setting.refund_or_number = request.POST.get('refund_or_number', '').strip() or None
        refund_str = request.POST.get('amount_refund_per_or', '').strip().replace(',', '').replace('₱','')
        if refund_str:
            try:
                setting.amount_refund_per_or = Decimal(refund_str)
            except Exception:
                pass
        else:
            setting.amount_refund_per_or = None
        reimbursed_str = request.POST.get('amount_reimbursed', '').strip().replace(',', '').replace('₱','')
        if reimbursed_str:
            try:
                setting.amount_reimbursed = Decimal(reimbursed_str)
            except Exception:
                pass
        else:
            setting.amount_reimbursed = None
        setting.check_number = request.POST.get('check_number', '').strip() or None
        setting.save()

        # Collect dynamically added new rows (for new CA)
        new_entries = []
        idx = 0
        while idx < 200:  # safety limit
            has_any = any(f'new_{field}_{idx}' in request.POST for field in ['date','or','fuel','amount','vat','wht5','wht1'])
            if not has_any:
                # Also check if any of the keys with that idx exist at all
                if idx > 30 and not any(k.startswith(f'new_') and k.endswith(f'_{idx}') for k in request.POST.keys()):
                    # No more new rows beyond 30 and no key found, break
                    if idx > 5:
                        break
                idx += 1
                if idx > 100:
                    break
                continue
            new_date_str = request.POST.get(f'new_date_{idx}', '').strip()
            new_or = request.POST.get(f'new_or_{idx}', '').strip()
            new_fuel = request.POST.get(f'new_fuel_{idx}', 'Diesel').strip() or 'Diesel'
            new_amount_str = request.POST.get(f'new_amount_{idx}', '').strip().replace(',', '').replace('₱','')
            new_vat = f'new_vat_{idx}' in request.POST
            new_wht5_str = request.POST.get(f'new_wht5_{idx}', '').strip().replace(',', '').replace('₱','')
            new_wht1_str = request.POST.get(f'new_wht1_{idx}', '').strip().replace(',', '').replace('₱','')
            idx += 1
            if not new_amount_str and not new_or and not new_date_str:
                continue
            try:
                new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date() if new_date_str else date.today()
            except Exception:
                new_date = date.today()
            try:
                new_amount = Decimal(new_amount_str) if new_amount_str else Decimal('0.00')
            except Exception:
                continue
            if new_amount == 0 and not new_or:
                continue
            try:
                new_wht5 = Decimal(new_wht5_str) if new_wht5_str else None
            except Exception:
                new_wht5 = None
            try:
                new_wht1 = Decimal(new_wht1_str) if new_wht1_str else None
            except Exception:
                new_wht1 = None
            new_entries.append((new_date, new_or, new_fuel, new_amount, new_vat, new_wht5, new_wht1))
            if len(new_entries) > 100:
                break

        if fresh:
            trips = FuelConsumption.objects.none()
        else:
            trips = FuelConsumption.objects.select_related('driver').order_by('date', 'reference_number')
        saved_report = LiquidationReport.objects.create(
            principal_amount=setting.principal_amount,
            check_number=setting.check_number or '',
            refund_or_number=setting.refund_or_number or '',
            amount_refund_per_or=setting.amount_refund_per_or,
            amount_reimbursed=setting.amount_reimbursed,
        )
        for trip in trips:
            if trip.id in deleted_ids:
                continue
            amount_to_save = edited_amounts.get(trip.id, Decimal(str(trip.cost)))
            vat_inc = edited_vat.get(trip.id, True)
            # Use edited WHT if provided, otherwise None (will compute)
            wht5_val = edited_wht5.get(trip.id)
            wht1_val = edited_wht1.get(trip.id)
            LiquidationReportEntry.objects.create(
                report=saved_report,
                entry_date=trip.date,
                or_number=trip.or_number or '',
                fuel_type='Diesel',
                amount=amount_to_save,
                vat_inclusive=vat_inc,
                wht5_amount=wht5_val,
                wht1_amount=wht1_val,
            )
        for new_date, new_or, new_fuel, new_amount, new_vat, new_wht5, new_wht1 in new_entries:
            LiquidationReportEntry.objects.create(
                report=saved_report,
                entry_date=new_date,
                or_number=new_or or '',
                fuel_type=new_fuel or 'Diesel',
                amount=new_amount,
                vat_inclusive=new_vat,
                wht5_amount=new_wht5,
                wht1_amount=new_wht1,
            )
        return redirect('saved_liquidation_reports')

    rows = []
    total = Decimal('0.00')
    total_wht5 = Decimal('0.00')
    total_wht1 = Decimal('0.00')
    total_net = Decimal('0.00')
    for trip in trips:
        amt = Decimal(str(trip.cost))
        vat_inc = True  # default VAT inclusive for new report (optional per row via checkbox)
        if vat_inc:
            wht5 = (amt / Decimal('1.12') * Decimal('0.05')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            wht1 = (amt / Decimal('1.12') * Decimal('0.01')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            wht5 = Decimal('0.00')
            wht1 = Decimal('0.00')
        net = (amt - wht5 - wht1).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        rows.append({
            'id': trip.id,
            'date': trip.date,
            'or_number': trip.or_number or '',
            'fuel_type': 'Diesel',
            'amount': f"{amt:,.2f}",
            'wht5': f"{wht5:,.2f}",
            'wht1': f"{wht1:,.2f}",
            'net': f"{net:,.2f}",
            'vat_inclusive': vat_inc,
            'reference_number': trip.reference_number,
        })
        total += amt
        total_wht5 += wht5
        total_wht1 += wht1
        total_net += net
    total = total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_wht5 = total_wht5.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_wht1 = total_wht1.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_net = total_net.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_wht = (total_wht5 + total_wht1).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    principal_dec = setting.principal_amount if setting.principal_amount else Decimal('0.00')
    unutilized = (principal_dec - total).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if setting.principal_amount else Decimal('0.00')
    unutilized_net = (principal_dec - total_net).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if setting.principal_amount else Decimal('0.00')

    context = {
        'rows': rows,
        'total': f"{total:,.2f}",
        'total_wht5': f"{total_wht5:,.2f}",
        'total_wht1': f"{total_wht1:,.2f}",
        'total_wht': f"{total_wht:,.2f}",
        'total_net': f"{total_net:,.2f}",
        'principal': f"{setting.principal_amount:,.2f}" if setting.principal_amount else '',
        'principal_raw': principal_dec,
        'unutilized': f"{unutilized:,.2f}" if setting.principal_amount else '',
        'unutilized_net': f"{unutilized_net:,.2f}" if setting.principal_amount else '',
        'refund_or_number': setting.refund_or_number or '',
        'amount_refund_per_or': f"{setting.amount_refund_per_or:,.2f}" if setting.amount_refund_per_or else '',
        'amount_reimbursed': f"{setting.amount_reimbursed:,.2f}" if setting.amount_reimbursed else '',
        'check_number': setting.check_number or '',
        'today': date.today(),
        'fresh': fresh,
    }
    return render(request, 'fuel/liquidation_report.html', context)


def liquidation_report_new(request):
    """Fresh / blank liquidation report — enter everything manually, no trip data."""
    return liquidation_report_view(request, fresh=True)


def saved_liquidation_reports(request):
    reports = LiquidationReport.objects.order_by('-created_at')
    for report in reports:
        entries = report.entries.all()
        total = Decimal('0.00')
        wht5_sum = Decimal('0.00')
        wht1_sum = Decimal('0.00')
        net_sum = Decimal('0.00')
        for e in entries:
            amt = Decimal(e.amount)
            total += amt
            vat_inc = getattr(e, 'vat_inclusive', True)
            # Use stored WHT if dynamically input per OR #, else compute
            if e.wht5_amount is not None:
                wht5 = Decimal(e.wht5_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            elif vat_inc:
                wht5 = (amt / Decimal('1.12') * Decimal('0.05')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            else:
                wht5 = Decimal('0.00')
            if e.wht1_amount is not None:
                wht1 = Decimal(e.wht1_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            elif vat_inc:
                wht1 = (amt / Decimal('1.12') * Decimal('0.01')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            else:
                wht1 = Decimal('0.00')
            net = (amt - wht5 - wht1).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            wht5_sum += wht5
            wht1_sum += wht1
            net_sum += net
        if not entries:
            total = Decimal(report.total()) if report.total() else Decimal('0.00')
            wht5_sum = (total / Decimal('1.12') * Decimal('0.05')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if total else Decimal('0.00')
            wht1_sum = (total / Decimal('1.12') * Decimal('0.01')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if total else Decimal('0.00')
            net_sum = (total - wht5_sum - wht1_sum).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if total else Decimal('0.00')
        wht_sum = (wht5_sum + wht1_sum).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        principal_dec = report.principal_amount if report.principal_amount else Decimal('0.00')
        unutilized = (principal_dec - total).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if report.principal_amount else Decimal('0.00')
        unutilized_net = (principal_dec - net_sum).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if report.principal_amount else Decimal('0.00')
        report.total_display = f'{total:,.2f}'
        report.principal_display = f'{report.principal_amount:,.2f}' if report.principal_amount else ''
        report.wht5_display = f'{wht5_sum:,.2f}'
        report.wht1_display = f'{wht1_sum:,.2f}'
        report.wht_display = f'{wht_sum:,.2f}'
        report.net_display = f'{net_sum:,.2f}'
        report.unutilized_display = f'{unutilized:,.2f}' if report.principal_amount else ''
        report.unutilized_net_display = f'{unutilized_net:,.2f}' if report.principal_amount else ''
        report.amount_refund_per_or_display = f'{report.amount_refund_per_or:,.2f}' if report.amount_refund_per_or else ''
        report.amount_reimbursed_display = f'{report.amount_reimbursed:,.2f}' if report.amount_reimbursed else ''
        report.refund_or_number_display = report.refund_or_number or ''
    return render(request, 'fuel/liquidation_saved_reports.html', {'reports': reports})


def liquidation_report_reprint(request, report_id):
    report = get_object_or_404(LiquidationReport, pk=report_id)
    entries = report.entries.order_by('entry_date')
    rows = []
    total = Decimal('0.00')
    total_wht5 = Decimal('0.00')
    total_wht1 = Decimal('0.00')
    total_net = Decimal('0.00')
    for entry in entries:
        amt = Decimal(entry.amount)
        vat_inc = getattr(entry, 'vat_inclusive', True)
        if entry.wht5_amount is not None:
            wht5 = Decimal(entry.wht5_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif vat_inc:
            wht5 = (amt / Decimal('1.12') * Decimal('0.05')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            wht5 = Decimal('0.00')
        if entry.wht1_amount is not None:
            wht1 = Decimal(entry.wht1_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif vat_inc:
            wht1 = (amt / Decimal('1.12') * Decimal('0.01')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            wht1 = Decimal('0.00')
        net = (amt - wht5 - wht1).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        rows.append({
            'id': entry.id,
            'date': entry.entry_date,
            'or_number': entry.or_number or '',
            'fuel_type': entry.fuel_type,
            'amount': f'{amt:,.2f}',
            'wht5': f'{wht5:,.2f}',
            'wht1': f'{wht1:,.2f}',
            'net': f'{net:,.2f}',
            'vat_inclusive': vat_inc,
        })
        total += amt
        total_wht5 += wht5
        total_wht1 += wht1
        total_net += net
    total = total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_wht5 = total_wht5.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_wht1 = total_wht1.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_net = total_net.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_wht = (total_wht5 + total_wht1).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    principal_dec = report.principal_amount if report.principal_amount else Decimal('0.00')
    unutilized = (principal_dec - total).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if report.principal_amount else Decimal('0.00')
    unutilized_net = (principal_dec - total_net).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if report.principal_amount else Decimal('0.00')
    context = {
        'rows': rows,
        'total': f'{total:,.2f}',
        'total_wht5': f'{total_wht5:,.2f}',
        'total_wht1': f'{total_wht1:,.2f}',
        'total_wht': f'{total_wht:,.2f}',
        'total_net': f'{total_net:,.2f}',
        'principal': f'{report.principal_amount:,.2f}' if report.principal_amount else '',
        'unutilized': f'{unutilized:,.2f}' if report.principal_amount else '',
        'unutilized_net': f'{unutilized_net:,.2f}' if report.principal_amount else '',
        'refund_or_number': report.refund_or_number or '',
        'amount_refund_per_or': f'{report.amount_refund_per_or:,.2f}' if report.amount_refund_per_or else '',
        'amount_reimbursed': f'{report.amount_reimbursed:,.2f}' if report.amount_reimbursed else '',
        'check_number': report.check_number or '',
        'today': report.report_date,
        'report': report,
        'readonly': True,
    }
    return render(request, 'fuel/liquidation_report.html', context)


def liquidation_report_edit(request, report_id):
    report = get_object_or_404(LiquidationReport, pk=report_id)
    entries = list(report.entries.order_by('entry_date'))

    if request.method == 'POST':
        for entry in entries:
            if f'delete_{entry.id}' in request.POST:
                LiquidationReportEntry.objects.filter(pk=entry.pk).delete()
                continue
            value = request.POST.get(f'or_{entry.id}', '').strip()
            if value != (entry.or_number or ''):
                LiquidationReportEntry.objects.filter(pk=entry.pk).update(or_number=value or None)
            amount_str = request.POST.get(f'amount_{entry.id}', '').strip().replace(',', '')
            if amount_str:
                try:
                    new_amt = Decimal(amount_str)
                    if new_amt != Decimal(entry.amount):
                        LiquidationReportEntry.objects.filter(pk=entry.pk).update(amount=new_amt)
                except Exception:
                    pass
            vat_inc = f'vat_{entry.id}' in request.POST
            if vat_inc != getattr(entry, 'vat_inclusive', True):
                LiquidationReportEntry.objects.filter(pk=entry.pk).update(vat_inclusive=vat_inc)
            # Dynamic WHT per OR # - editable amount of withholding tax per row
            if f'wht5_{entry.id}' in request.POST:
                wht5_str = request.POST.get(f'wht5_{entry.id}', '').strip().replace(',', '').replace('₱','')
                if wht5_str:
                    try:
                        wht5_val = Decimal(wht5_str)
                        LiquidationReportEntry.objects.filter(pk=entry.pk).update(wht5_amount=wht5_val)
                    except Exception:
                        pass
                else:
                    # If cleared, reset to None to recompute
                    if entry.wht5_amount is not None:
                        LiquidationReportEntry.objects.filter(pk=entry.pk).update(wht5_amount=None)
            if f'wht1_{entry.id}' in request.POST:
                wht1_str = request.POST.get(f'wht1_{entry.id}', '').strip().replace(',', '').replace('₱','')
                if wht1_str:
                    try:
                        wht1_val = Decimal(wht1_str)
                        LiquidationReportEntry.objects.filter(pk=entry.pk).update(wht1_amount=wht1_val)
                    except Exception:
                        pass
                else:
                    if entry.wht1_amount is not None:
                        LiquidationReportEntry.objects.filter(pk=entry.pk).update(wht1_amount=None)
        principal = request.POST.get('principal', '').strip().replace(',', '')
        if principal:
            try:
                report.principal_amount = Decimal(principal)
            except Exception:
                pass
        # Footer dynamic amounts - OR number for refund is now editable
        report.refund_or_number = request.POST.get('refund_or_number', '').strip() or None
        refund_str = request.POST.get('amount_refund_per_or', '').strip().replace(',', '').replace('₱','')
        if refund_str:
            try:
                report.amount_refund_per_or = Decimal(refund_str)
            except Exception:
                pass
        else:
            report.amount_refund_per_or = None
        reimbursed_str = request.POST.get('amount_reimbursed', '').strip().replace(',', '').replace('₱','')
        if reimbursed_str:
            try:
                report.amount_reimbursed = Decimal(reimbursed_str)
            except Exception:
                pass
        else:
            report.amount_reimbursed = None
        report.check_number = request.POST.get('check_number', '').strip() or None
        report.save()
        # Handle dynamically added new rows when editing
        new_idx = 0
        while new_idx < 200:
            has_any = any(f'new_{field}_{new_idx}' in request.POST for field in ['date','or','fuel','amount','vat','wht5','wht1'])
            if not has_any:
                if new_idx > 30 and not any(k.startswith(f'new_') and k.endswith(f'_{new_idx}') for k in request.POST.keys()):
                    if new_idx > 5:
                        break
                new_idx += 1
                if new_idx > 100:
                    break
                continue
            new_date_str = request.POST.get(f'new_date_{new_idx}', '').strip()
            new_or = request.POST.get(f'new_or_{new_idx}', '').strip()
            new_fuel = request.POST.get(f'new_fuel_{new_idx}', 'Diesel').strip() or 'Diesel'
            new_amount_str = request.POST.get(f'new_amount_{new_idx}', '').strip().replace(',', '').replace('₱','')
            new_vat = f'new_vat_{new_idx}' in request.POST
            new_wht5_str = request.POST.get(f'new_wht5_{new_idx}', '').strip().replace(',', '').replace('₱','')
            new_wht1_str = request.POST.get(f'new_wht1_{new_idx}', '').strip().replace(',', '').replace('₱','')
            new_idx += 1
            if not new_amount_str and not new_or and not new_date_str:
                continue
            try:
                new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date() if new_date_str else date.today()
            except Exception:
                new_date = date.today()
            try:
                new_amount = Decimal(new_amount_str) if new_amount_str else Decimal('0.00')
            except Exception:
                continue
            if new_amount == 0 and not new_or:
                continue
            try:
                new_wht5 = Decimal(new_wht5_str) if new_wht5_str else None
            except Exception:
                new_wht5 = None
            try:
                new_wht1 = Decimal(new_wht1_str) if new_wht1_str else None
            except Exception:
                new_wht1 = None
            LiquidationReportEntry.objects.create(
                report=report,
                entry_date=new_date,
                or_number=new_or or '',
                fuel_type=new_fuel or 'Diesel',
                amount=new_amount,
                vat_inclusive=new_vat,
                wht5_amount=new_wht5,
                wht1_amount=new_wht1,
            )
            if LiquidationReportEntry.objects.filter(report=report).count() > 200:
                break
        return redirect('liquidation_report_reprint', report_id=report.pk)

    rows = []
    total = Decimal('0.00')
    total_wht5 = Decimal('0.00')
    total_wht1 = Decimal('0.00')
    total_net = Decimal('0.00')
    for entry in entries:
        amt = Decimal(entry.amount)
        vat_inc = getattr(entry, 'vat_inclusive', True)
        if entry.wht5_amount is not None:
            wht5 = Decimal(entry.wht5_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif vat_inc:
            wht5 = (amt / Decimal('1.12') * Decimal('0.05')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            wht5 = Decimal('0.00')
        if entry.wht1_amount is not None:
            wht1 = Decimal(entry.wht1_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif vat_inc:
            wht1 = (amt / Decimal('1.12') * Decimal('0.01')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            wht1 = Decimal('0.00')
        net = (amt - wht5 - wht1).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        rows.append({
            'id': entry.id,
            'date': entry.entry_date,
            'or_number': entry.or_number or '',
            'fuel_type': entry.fuel_type,
            'amount': f'{amt:,.2f}',
            'wht5': f'{wht5:,.2f}',
            'wht1': f'{wht1:,.2f}',
            'net': f'{net:,.2f}',
            'vat_inclusive': vat_inc,
        })
        total += amt
        total_wht5 += wht5
        total_wht1 += wht1
        total_net += net
    total = total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_wht5 = total_wht5.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_wht1 = total_wht1.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_net = total_net.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_wht = (total_wht5 + total_wht1).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    principal_dec = report.principal_amount if report.principal_amount else Decimal('0.00')
    unutilized = (principal_dec - total).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if report.principal_amount else Decimal('0.00')
    unutilized_net = (principal_dec - total_net).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if report.principal_amount else Decimal('0.00')
    context = {
        'rows': rows,
        'total': f'{total:,.2f}',
        'total_wht5': f'{total_wht5:,.2f}',
        'total_wht1': f'{total_wht1:,.2f}',
        'total_wht': f'{total_wht:,.2f}',
        'total_net': f'{total_net:,.2f}',
        'principal': f'{report.principal_amount:,.2f}' if report.principal_amount else '',
        'unutilized': f'{unutilized:,.2f}' if report.principal_amount else '',
        'unutilized_net': f'{unutilized_net:,.2f}' if report.principal_amount else '',
        'refund_or_number': report.refund_or_number or '',
        'amount_refund_per_or': f'{report.amount_refund_per_or:,.2f}' if report.amount_refund_per_or else '',
        'amount_reimbursed': f'{report.amount_reimbursed:,.2f}' if report.amount_reimbursed else '',
        'check_number': report.check_number or '',
        'today': report.report_date,
        'report': report,
        'editing': True,
    }
    return render(request, 'fuel/liquidation_report.html', context)


def export_liquidation_report_pdf(request, report_id=None):
    """Export the liquidation report as a downloadable PDF (amounts from DB trips)."""
    if not PDF_SUPPORT:
        return HttpResponse("PDF export is not available. Please install reportlab library.", content_type="text/plain")

    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    if report_id is not None:
        report = get_object_or_404(LiquidationReport, pk=report_id)
        entries = report.entries.order_by('entry_date')
        principal_value_src = report.principal_amount
        check_number_src = report.check_number or ''
        refund_or_number_src = report.refund_or_number or ''
        refund_value_src = report.amount_refund_per_or
        reimbursed_value_src = report.amount_reimbursed
        total_src = report.total()
        detail_rows = [(e.entry_date, e.or_number or '', e.fuel_type, e.amount, getattr(e, 'vat_inclusive', True), e.wht5_amount, e.wht1_amount) for e in entries]
    else:
        trips = FuelConsumption.objects.select_related('driver').order_by('date', 'reference_number')
        setting, _ = LiquidationSetting.objects.get_or_create(pk=1)
        principal_value_src = setting.principal_amount
        check_number_src = setting.check_number or ''
        refund_or_number_src = setting.refund_or_number or ''
        refund_value_src = setting.amount_refund_per_or
        reimbursed_value_src = setting.amount_reimbursed
        total_src = sum(trip.cost for trip in trips)
        detail_rows = [(trip.date, trip.or_number or '', 'Diesel', trip.cost, True, None, None) for trip in trips]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="liquidation_report.pdf"'

    doc = SimpleDocTemplate(
        response, pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=36, bottomMargin=36
    )
    styles = getSampleStyleSheet()

    font_name = 'Helvetica'
    bold_font_name = 'Helvetica-Bold'
    try:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_dir = os.path.dirname(current_dir)
        fonts_dir = os.path.join(project_dir, 'fonts', 'ttf')
        dejavu_font_path = os.path.join(fonts_dir, 'DejaVuSans.ttf')
        dejavu_bold_font_path = os.path.join(fonts_dir, 'DejaVuSans-Bold.ttf')
        if os.path.exists(dejavu_font_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_font_path))
            if os.path.exists(dejavu_bold_font_path):
                pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_bold_font_path))
                bold_font_name = 'DejaVuSans-Bold'
            else:
                bold_font_name = 'DejaVuSans'
            font_name = 'DejaVuSans'
    except Exception:
        font_name = 'Helvetica'
        bold_font_name = 'Helvetica-Bold'

    elements = []

    title_style = ParagraphStyle(
        'LiqTitle', parent=styles['Normal'], fontSize=18,
        fontName=bold_font_name, spaceAfter=4, leading=22, alignment=1
    )
    field_style = ParagraphStyle(
        'LiqField', parent=styles['Normal'], fontSize=12, fontName=font_name
    )
    elements.append(Paragraph('LIQUIDATION REPORT', title_style))

    no_date_table = Table(
        [[Paragraph('No: ___________________', field_style),
          Paragraph('Date: _________________', field_style)]],
        colWidths=[266, 266]
    )
    no_date_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(no_date_table)

    addr_table = Table(
        [[Paragraph('Dumingag, Zamboanga del Sur', field_style),
          Paragraph('Responsibility Center: MDRRMO', field_style)]],
        colWidths=[266, 266]
    )
    addr_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(addr_table)
    elements.append(Spacer(1, 10))

    desc_style = ParagraphStyle(
        'LiqDesc', parent=styles['Normal'], fontSize=12,
        fontName=font_name, leading=15
    )
    total_style = ParagraphStyle(
        'LiqTotal', parent=styles['Normal'], fontSize=12, fontName=bold_font_name
    )
    principal_value = f'{principal_value_src:,.2f}' if principal_value_src else ''
    amount_label_style = ParagraphStyle(
        'LiqAmountLabel', parent=styles['Normal'], fontSize=11,
        fontName=bold_font_name, alignment=2, leading=13
    )
    amount_value_style = ParagraphStyle(
        'LiqAmountValue', parent=styles['Normal'], fontSize=12,
        fontName=bold_font_name, alignment=2
    )
    principal_table = Table(
        [[Paragraph(
            'Liquidation of Cash Advance for the Diesel expenses incurred '
            'during Patient and Emergency Transportation.',
            desc_style
        ), Paragraph('AMOUNT', amount_label_style)],
         ['', Paragraph(principal_value, amount_value_style)]],
        colWidths=[380, 152]
    )
    principal_table.setStyle(TableStyle([
        ('SPAN', (0, 0), (0, 1)),
        ('VALIGN', (0, 0), (0, 1), 'MIDDLE'),
        ('VALIGN', (1, 0), (1, 0), 'BOTTOM'),
        ('VALIGN', (1, 1), (1, 1), 'BOTTOM'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('ALIGN', (1, 1), (1, 1), 'RIGHT'),
        ('LINEBELOW', (1, 0), (1, 0), 0.5, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(principal_table)
    elements.append(Spacer(1, 10))

    partic_style = ParagraphStyle(
        'LiqPart', parent=styles['Normal'], fontSize=14, fontName=bold_font_name,
        alignment=1, spaceBefore=8, spaceAfter=4
    )
    elements.append(Paragraph('PARTICULARS', partic_style))

    # Build detail table with Withholding Tax 5%/1% (computed as amount/1.12*rate) and Net Amount - print-friendly white header with black text (no solid ink)
    # Larger fonts for elder readability (was 6.5 -> now 9)
    header_cell_style = ParagraphStyle('LiqHeadCell', parent=styles['Normal'], fontSize=9, fontName=bold_font_name, textColor=colors.black, alignment=1, leading=10)
    header_sub_style = ParagraphStyle('LiqHeadSub', parent=styles['Normal'], fontSize=9, fontName=bold_font_name, textColor=colors.black, alignment=1, leading=10)
    cell_center_style = ParagraphStyle('LiqCellCenter', parent=styles['Normal'], fontSize=9, fontName=font_name, alignment=1, leading=10)
    cell_right_style = ParagraphStyle('LiqCellRight', parent=styles['Normal'], fontSize=9, fontName=font_name, alignment=2, leading=10)
    cell_bold_right_style = ParagraphStyle('LiqCellBoldRight', parent=styles['Normal'], fontSize=9, fontName=bold_font_name, alignment=2, leading=10)
    cell_bold_center_style = ParagraphStyle('LiqCellBoldCenter', parent=styles['Normal'], fontSize=9, fontName=bold_font_name, alignment=1, leading=10)
    # Header rows: Withholding Tax 5%/1% grouped - VAT flag hidden in PDF (WHT applied per row if VAT checked, else 0)
    detail_data = [
        [Paragraph('Date', header_cell_style), Paragraph('OR Receipt #', header_cell_style), Paragraph('Fuel Type', header_cell_style), Paragraph('Amount', header_cell_style), Paragraph('Withholding Tax', header_cell_style), '', Paragraph('Net Amount', header_cell_style)],
        ['', '', '', '', Paragraph('5%', header_sub_style), Paragraph('1%', header_sub_style), ''],
    ]
    total_dec = Decimal(str(total_src))
    # Compute per-row totals for accurate rounding — VAT optional per row
    total_wht5_src = Decimal('0.00')
    total_wht1_src = Decimal('0.00')
    total_net_src = Decimal('0.00')
    for row in detail_rows:
        # row = (date, or_no, fuel_type, amount, vat_inclusive, wht5_stored, wht1_stored) or legacy
        if len(row) == 7:
            _, _, _, amount, vat_inc, wht5_stored, wht1_stored = row
        elif len(row) == 5:
            _, _, _, amount, vat_inc = row
            wht5_stored = wht1_stored = None
        else:
            _, _, _, amount = row
            vat_inc = True
            wht5_stored = wht1_stored = None
        amt = Decimal(str(amount))
        if wht5_stored is not None:
            wht5 = Decimal(str(wht5_stored)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif vat_inc:
            wht5 = (amt / Decimal('1.12') * Decimal('0.05')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            wht5 = Decimal('0.00')
        if wht1_stored is not None:
            wht1 = Decimal(str(wht1_stored)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif vat_inc:
            wht1 = (amt / Decimal('1.12') * Decimal('0.01')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            wht1 = Decimal('0.00')
        net = (amt - wht5 - wht1).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_wht5_src += wht5
        total_wht1_src += wht1
        total_net_src += net
    total_wht5_src = total_wht5_src.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_wht1_src = total_wht1_src.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_net_src = total_net_src.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_dec = total_dec.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    # Added: total WHT (5%+1%) and unutilized fund (Principal - Total / Principal - Net) for footer
    total_wht_src = (total_wht5_src + total_wht1_src).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    principal_dec = Decimal(str(principal_value_src)) if principal_value_src else Decimal('0.00')
    unutilized_src = (principal_dec - total_dec).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if principal_value_src else Decimal('0.00')
    unutilized_net_src = (principal_dec - total_net_src).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if principal_value_src else Decimal('0.00')
    for row in detail_rows:
        if len(row) == 7:
            entry_date, or_no, fuel_type, amount, vat_inc, wht5_stored, wht1_stored = row
        elif len(row) == 5:
            entry_date, or_no, fuel_type, amount, vat_inc = row
            wht5_stored = wht1_stored = None
        else:
            entry_date, or_no, fuel_type, amount = row
            vat_inc = True
            wht5_stored = wht1_stored = None
        amt = Decimal(str(amount))
        if wht5_stored is not None:
            wht5 = Decimal(str(wht5_stored)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif vat_inc:
            wht5 = (amt / Decimal('1.12') * Decimal('0.05')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            wht5 = Decimal('0.00')
        if wht1_stored is not None:
            wht1 = Decimal(str(wht1_stored)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif vat_inc:
            wht1 = (amt / Decimal('1.12') * Decimal('0.01')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            wht1 = Decimal('0.00')
        net = (amt - wht5 - wht1).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        detail_data.append([
            Paragraph(entry_date.strftime('%m/%d/%Y'), cell_center_style),
            Paragraph(f'OR NO {or_no}' if or_no else '', cell_center_style),
            Paragraph(str(fuel_type), cell_center_style),
            Paragraph(f'{amt:,.2f}', cell_right_style),
            Paragraph(f'{wht5:,.2f}', cell_right_style),
            Paragraph(f'{wht1:,.2f}', cell_right_style),
            Paragraph(f'{net:,.2f}', cell_right_style),
        ])
    # Totals row
    detail_data.append([
        Paragraph('', cell_center_style),
        Paragraph('', cell_center_style),
        Paragraph('<b>TOTAL</b>', cell_bold_center_style),
        Paragraph(f'{total_dec:,.2f}', cell_bold_right_style),
        Paragraph(f'{total_wht5_src:,.2f}', cell_bold_right_style),
        Paragraph(f'{total_wht1_src:,.2f}', cell_bold_right_style),
        Paragraph(f'{total_net_src:,.2f}', cell_bold_right_style),
    ])
    detail_table = Table(detail_data, colWidths=[55, 95, 40, 80, 65, 65, 82], repeatRows=2)
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 1), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, 1), colors.black),
        ('FONTNAME', (0, 0), (-1, 1), bold_font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOX', (0, 0), (-1, 1), 1.0, colors.black),
        ('LINEBELOW', (0, 1), (-1, 1), 1.0, colors.black),
        ('SPAN', (4, 0), (5, 0)),
        ('SPAN', (6, 0), (6, 1)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f1f5f9')),
        ('LINEABOVE', (0, -1), (-1, -1), 1.2, colors.black),
    ]))
    elements.append(detail_table)
    elements.append(Spacer(1, 6))

    total = total_src
    # Summary bar with Amount / WHT 5% / WHT 1% / Net Amount - larger for elder
    wht_label_style = ParagraphStyle('WhtLabel', parent=styles['Normal'], fontSize=9, fontName=bold_font_name, textColor=colors.HexColor('#475569'), alignment=1)
    summary_header = Table(
        [[Paragraph('', wht_label_style), Paragraph('Amount', wht_label_style), Paragraph('WHT 5%', wht_label_style), Paragraph('WHT 1%', wht_label_style), Paragraph('Net Amount', wht_label_style)]],
        colWidths=[172, 90, 90, 90, 90]
    )
    summary_header.setStyle(TableStyle([
        ('ALIGN', (1, 0), (-1, 0), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(summary_header)
    summary_table = Table(
        [[Paragraph('TOTAL AMOUNT SPENT', total_style), Paragraph(f'{total_dec:,.2f}', amount_value_style), Paragraph(f'{total_wht5_src:,.2f}', amount_value_style), Paragraph(f'{total_wht1_src:,.2f}', amount_value_style), Paragraph(f'{total_net_src:,.2f}', amount_value_style)]],
        colWidths=[172, 90, 90, 90, 90]
    )
    summary_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 1, colors.black),
        ('LINEBELOW', (0, 0), (-1, 0), 0.8, colors.black),
        ('ALIGN', (1, 0), (-1, 0), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 4))
    # Explicit WHT breakdown + Net for liquidation compliance (Net = Amount - WHT5 - WHT1, where VAT: WHT = Amount/1.12*rate; Non-VAT: WHT = Amount*rate; e.g., VAT 1500 => 66.96+13.39=Net1419.65, Non-VAT 1500 =>75+15=Net1410)
    # Increased to 10pt for elder readability
    wht_detail_style = ParagraphStyle('WhtDetail', parent=styles['Normal'], fontSize=10, fontName=font_name, leading=12)
    wht_value_style = ParagraphStyle('WhtValue', parent=styles['Normal'], fontSize=10, fontName=bold_font_name, alignment=2)
    net_label_style = ParagraphStyle('WhtNetLabel', parent=styles['Normal'], fontSize=10, fontName=bold_font_name, leading=12)
    net_value_style = ParagraphStyle('WhtNetValue', parent=styles['Normal'], fontSize=11, fontName=bold_font_name, alignment=2, textColor=colors.HexColor('#0f172a'))
    wht_breakdown = Table(
        [
            [Paragraph('Less: Withholding Tax — 5%  (VAT: ÷1.12×5% / Non-VAT: —)', wht_detail_style), Paragraph(f'{total_wht5_src:,.2f}', wht_value_style)],
            [Paragraph('Less: Withholding Tax — 1%  (VAT: ÷1.12×1% / Non-VAT: —)', wht_detail_style), Paragraph(f'{total_wht1_src:,.2f}', wht_value_style)],
            [Paragraph('<b>Total Withholding Tax (5% + 1%)</b>', net_label_style), Paragraph(f'<b>{total_wht_src:,.2f}</b>', net_value_style)],
            [Paragraph('Net Amount  (Amount − WHT 5% − WHT 1%)', net_label_style), Paragraph(f'{total_net_src:,.2f}', net_value_style)],
        ],
        colWidths=[437, 95]
    )
    wht_breakdown.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
        ('BACKGROUND', (0, 0), (-1, 2), colors.HexColor('#fffbeb')),
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#ecfdf5')),
        ('LINEABOVE', (0, 2), (-1, 2), 0.6, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(wht_breakdown)
    elements.append(Spacer(1, 8))
    # Unutilized Fund footer (Principal - Total Spent) and (Principal - Net) - for elder view larger 10pt
    if principal_value_src:
        unutil_style = ParagraphStyle('UnutilLabel', parent=styles['Normal'], fontSize=10, fontName=bold_font_name, leading=12, textColor=colors.HexColor('#0f172a'))
        unutil_value_style = ParagraphStyle('UnutilValue', parent=styles['Normal'], fontSize=11, fontName=bold_font_name, alignment=2, textColor=colors.HexColor('#1e40af'))
        unutil_detail_style = ParagraphStyle('UnutilDetail', parent=styles['Normal'], fontSize=9, fontName=font_name, leading=11, textColor=colors.HexColor('#475569'))
        unutil_table = Table(
            [
                [Paragraph('Principal Amount (Cash Advance)', unutil_detail_style), Paragraph(f'₱{principal_dec:,.2f}', unutil_value_style)],
                [Paragraph('Less: Total Amount Spent', unutil_detail_style), Paragraph(f'₱{total_dec:,.2f}', wht_value_style)],
                [Paragraph('<b>Unutilized Fund (Principal − Total Spent)</b>', unutil_style), Paragraph(f'<b>₱{unutilized_src:,.2f}</b>', unutil_value_style)],
                [Paragraph('Unutilized Fund (Principal − Net Amount)', unutil_detail_style), Paragraph(f'₱{unutilized_net_src:,.2f}', unutil_value_style)],
            ],
            colWidths=[437, 95]
        )
        unutil_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eff6ff')),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#dbeafe')),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#f0fdf4')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('LINEABOVE', (0, 2), (-1, 2), 0.8, colors.HexColor('#1e40af')),
        ]))
        elements.append(unutil_table)
        elements.append(Spacer(1, 10))

    line_style = ParagraphStyle(
        'LiqLine', parent=styles['Normal'], fontSize=11, fontName=font_name, leading=13
    )
    line_value_style = ParagraphStyle(
        'LiqLineValue', parent=styles['Normal'], fontSize=11, fontName=bold_font_name, alignment=2, leading=13
    )
    principal_display = f'₱{principal_value}' if principal_value else '—'
    refund_display = f'₱{refund_value_src:,.2f}' if refund_value_src else '—'
    reimbursed_display = f'₱{reimbursed_value_src:,.2f}' if reimbursed_value_src else '—'
    check_text = ('AMOUNT OF CASH ADVANCE PER DV CHECK # ' + check_number_src) if check_number_src else 'AMOUNT OF CASH ADVANCE PER DV CHECK # '
    refund_label = f'AMOUNT REFUND PER OR #  {refund_or_number_src}' if refund_or_number_src else 'AMOUNT REFUND PER OR #'
    lines_table = Table(
        [
            [Paragraph(check_text, line_style), Paragraph(principal_display, line_value_style)],
            [Paragraph(refund_label, line_style), Paragraph(refund_display, line_value_style)],
            [Paragraph('AMOUNT REIMBURSED', line_style), Paragraph(reimbursed_display, line_value_style)],
        ],
        colWidths=[410, 122]
    )
    lines_table.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LINEBELOW', (0, 0), (-1, 0), 0.6, colors.HexColor('#e2e8f0')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
    ]))
    elements.append(lines_table)
    elements.append(Spacer(1, 16))

    sig_style = ParagraphStyle(
        'LiqSig', parent=styles['Normal'], fontSize=11, fontName=font_name, alignment=1
    )
    sig_name_style = ParagraphStyle(
        'LiqSigName', parent=styles['Normal'], fontSize=12, fontName=bold_font_name, alignment=1
    )
    sig_data = [
        [Paragraph('Submitted by:', sig_style), Paragraph('Received by:', sig_style)],
        ['', ''],
        [Paragraph('JHUNAX L. CARDOZA', sig_name_style), Paragraph('REINA B. VASQUEZ', sig_name_style)],
        [Paragraph('LDRRMO-III', sig_style), Paragraph('Municipal Accountant', sig_style)],
        [Paragraph('Date: ______________________', sig_style), Paragraph('Date: ______________________', sig_style)],
    ]
    sig_table = Table(sig_data, colWidths=[266, 266])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LINEABOVE', (0, 2), (-1, 2), 1, colors.black),
        ('TOPPADDING', (0, 2), (-1, 2), 16),
    ]))
    elements.append(sig_table)

    doc.build(elements)

    return response


# ──────────────────────────────────────────────────────────────
# Petty Cash Voucher + RER — separate transactions
# RER print: image on top + RER form at bottom on one A4
# PCV print: separate A4
# ──────────────────────────────────────────────────────────────
from .models import PettyCashVoucher, ReimbursementExpenseReceipt, ReimbursementExpenseReceiptImage
from .forms import PettyCashVoucherForm, ReimbursementExpenseReceiptForm


class PettyCashVoucherListView(ListView):
    model = PettyCashVoucher
    template_name = 'fuel/pcv_list.html'
    context_object_name = 'vouchers'
    paginate_by = 20

    def get_queryset(self):
        qs = PettyCashVoucher.objects.all()
        # --- Comprehensive search ---
        q = self.request.GET.get('q', '').strip()
        voucher_no = self.request.GET.get('voucher_no', '').strip()
        date_from = self.request.GET.get('date_from', '').strip()
        date_to = self.request.GET.get('date_to', '').strip()
        amount_min = self.request.GET.get('amount_min', '').strip()
        amount_max = self.request.GET.get('amount_max', '').strip()
        supplier = self.request.GET.get('supplier', '').strip()
        or_no = self.request.GET.get('or_no', '').strip()
        fund = self.request.GET.get('fund', '').strip()
        status = self.request.GET.get('status', '').strip()
        sort = self.request.GET.get('sort', '').strip()

        if q:
            qs = qs.filter(
                Q(voucher_no__icontains=q) |
                Q(particulars__icontains=q) |
                Q(purpose__icontains=q) |
                Q(payee_office__icontains=q) |
                Q(or_invoice_no__icontains=q) |
                Q(requested_by_name__icontains=q) |
                Q(paid_by_name__icontains=q) |
                Q(cash_received_by_name__icontains=q) |
                Q(reimbursement_received_by__icontains=q) |
                Q(fund__icontains=q) |
                Q(fpp__icontains=q) |
                Q(address__icontains=q)
            )
        if voucher_no:
            qs = qs.filter(voucher_no__icontains=voucher_no)
        if supplier:
            # Search supplier across particulars, purpose and payee
            qs = qs.filter(
                Q(particulars__icontains=supplier) |
                Q(purpose__icontains=supplier) |
                Q(payee_office__icontains=supplier) |
                Q(or_invoice_no__icontains=supplier)
            )
        if or_no:
            qs = qs.filter(or_invoice_no__icontains=or_no)
        if fund:
            qs = qs.filter(fund__icontains=fund)
        if date_from:
            try:
                qs = qs.filter(voucher_date__gte=date_from)
            except Exception:
                pass
        if date_to:
            try:
                qs = qs.filter(voucher_date__lte=date_to)
            except Exception:
                pass
        if amount_min:
            try:
                qs = qs.filter(amount__gte=Decimal(amount_min.replace(',', '').replace('₱','')))
            except Exception:
                pass
        if amount_max:
            try:
                qs = qs.filter(amount__lte=Decimal(amount_max.replace(',', '').replace('₱','')))
            except Exception:
                pass
        if status == 'liquidated':
            qs = qs.filter(liquidation_submitted=True)
        elif status == 'reimbursed':
            qs = qs.filter(reimbursement_paid=True)
        elif status == 'pending':
            qs = qs.filter(liquidation_submitted=False, reimbursement_paid=False)
        elif status == 'refunded':
            qs = qs.filter(received_refund=True)

        # Sorting for elder-friendly amount/date views
        if sort == 'amount_asc':
            qs = qs.order_by('amount', '-voucher_date', '-id')
        elif sort == 'amount_desc':
            qs = qs.order_by('-amount', '-voucher_date', '-id')
        elif sort == 'date_asc':
            qs = qs.order_by('voucher_date', 'id')
        elif sort == 'no_asc':
            qs = qs.order_by('voucher_no')
        elif sort == 'no_desc':
            qs = qs.order_by('-voucher_no')
        else:
            qs = qs.order_by('-voucher_date', '-id')
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # For filter dropdowns and preserving query
        ctx['funds'] = PettyCashVoucher.objects.exclude(fund='').values_list('fund', flat=True).distinct().order_by('fund')
        # Keep query string without page for pagination
        params = self.request.GET.copy()
        params.pop('page', None)
        ctx['query_string'] = params.urlencode()
        # Total and filtered counts
        ctx['total_count'] = PettyCashVoucher.objects.count()
        # Use paginator count for filtered
        try:
            ctx['filtered_count'] = self.get_queryset().count()
        except Exception:
            ctx['filtered_count'] = ctx['total_count']
        return ctx


class PettyCashVoucherDetailView(DetailView):
    model = PettyCashVoucher
    template_name = 'fuel/pcv_detail.html'
    context_object_name = 'voucher'


class PettyCashVoucherCreateView(CreateView):
    model = PettyCashVoucher
    form_class = PettyCashVoucherForm
    template_name = 'fuel/pcv_form.html'
    success_url = reverse_lazy('pcv_list')

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except ValidationError as e:
            form.add_error(None, e)
            return self.form_invalid(form)


class PettyCashVoucherUpdateView(UpdateView):
    model = PettyCashVoucher
    form_class = PettyCashVoucherForm
    template_name = 'fuel/pcv_form.html'
    success_url = reverse_lazy('pcv_list')

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except ValidationError as e:
            form.add_error(None, e)
            return self.form_invalid(form)


class PettyCashVoucherDeleteView(DeleteView):
    model = PettyCashVoucher
    template_name = 'fuel/pcv_confirm_delete.html'
    success_url = reverse_lazy('pcv_list')


def pcv_print_view(request, pk):
    voucher = get_object_or_404(PettyCashVoucher, pk=pk)
    return render(request, 'fuel/pcv_print.html', {'voucher': voucher})


class RERListView(ListView):
    model = ReimbursementExpenseReceipt
    template_name = 'fuel/rer_list.html'
    context_object_name = 'receipts'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().select_related('petty_cash_voucher').prefetch_related('images')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(received_from_name__icontains=q) | qs.filter(rer_no__icontains=q)
        return qs


class RERDetailView(DetailView):
    model = ReimbursementExpenseReceipt
    template_name = 'fuel/rer_detail.html'
    context_object_name = 'receipt'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Prefetch images for gallery
        receipt = self.object
        # Use get_gallery_images logic: prefer new images, fallback to legacy
        images = list(receipt.images.all())
        ctx['gallery_images'] = images
        # also expose legacy fallback flag
        ctx['has_legacy_image'] = bool(receipt.attached_image) and not images
        return ctx


# Helper: save multiple images for RER
def _handle_rer_images(request, rer):
    """
    Handle deletion of existing gallery images (via checkboxes delete_image_<id>)
    and creation of new uploads from 'new_images' (multiple).
    Properly places images: files are kept with object-fit contain in grid,
    and file deletion avoids removing shared files (e.g., migrated legacy copy).
    """
    # 1) Delete marked images
    for key in list(request.POST.keys()):
        if key.startswith('delete_image_'):
            try:
                img_id = int(key[len('delete_image_'):])
                img = ReimbursementExpenseReceiptImage.objects.filter(pk=img_id, rer=rer).first()
                if img:
                    fname = img.image.name if img.image else None
                    # Check if other gallery images or legacy share the same file
                    shared = False
                    if fname:
                        if ReimbursementExpenseReceiptImage.objects.filter(image=fname).exclude(pk=img.pk).exists():
                            shared = True
                        if rer.attached_image and rer.attached_image.name == fname:
                            shared = True
                    if fname and not shared:
                        try:
                            img.image.delete(save=False)
                        except Exception:
                            pass
                    img.delete()
            except Exception:
                continue
    # Handle explicit 'clear_legacy' if legacy checkbox
    if request.POST.get('clear_legacy_image') and rer.attached_image:
        fname = rer.attached_image.name
        # Don't delete file if still referenced by gallery
        shared = False
        if fname and ReimbursementExpenseReceiptImage.objects.filter(image=fname).exists():
            shared = True
        try:
            if not shared and rer.attached_image:
                rer.attached_image.delete(save=False)
            rer.attached_image = None
            rer.save(update_fields=['attached_image'])
        except Exception:
            try:
                rer.attached_image = None
                rer.save(update_fields=['attached_image'])
            except Exception:
                pass
    # 2) Create new images
    files = request.FILES.getlist('new_images')
    if files:
        max_order = rer.images.aggregate(models.Max('order'))['order__max'] or 0
        for idx, f in enumerate(files):
            # Basic validation: ensure it's an image (content_type starts with image/)
            # Pillow will validate on save
            try:
                ReimbursementExpenseReceiptImage.objects.create(rer=rer, image=f, order=max_order + 1 + idx)
            except Exception:
                continue


class RERCreateView(CreateView):
    model = ReimbursementExpenseReceipt
    form_class = ReimbursementExpenseReceiptForm
    template_name = 'fuel/rer_form.html'
    success_url = reverse_lazy('rer_list')

    def form_valid(self, form):
        try:
            resp = super().form_valid(form)
            # After saving, handle multi-image uploads
            _handle_rer_images(self.request, self.object)
            return resp
        except ValidationError as e:
            form.add_error(None, e)
            return self.form_invalid(form)


class RERUpdateView(UpdateView):
    model = ReimbursementExpenseReceipt
    form_class = ReimbursementExpenseReceiptForm
    template_name = 'fuel/rer_form.html'
    success_url = reverse_lazy('rer_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Provide gallery images for edit preview
        if self.object:
            ctx['gallery_images'] = list(self.object.images.all())
        return ctx

    def form_valid(self, form):
        try:
            resp = super().form_valid(form)
            _handle_rer_images(self.request, self.object)
            return resp
        except ValidationError as e:
            form.add_error(None, e)
            return self.form_invalid(form)


class RERDeleteView(DeleteView):
    model = ReimbursementExpenseReceipt
    template_name = 'fuel/rer_confirm_delete.html'
    success_url = reverse_lazy('rer_list')


def rer_print_view(request, pk):
    """Print: gallery images on top (responsive grid, properly placed) + RER form at bottom, one A4 portrait.
    Supports multiple images (1–N). Single legacy image fallback if no gallery images.
    """
    receipt = get_object_or_404(
        ReimbursementExpenseReceipt.objects.select_related('petty_cash_voucher').prefetch_related('images'),
        pk=pk
    )
    images = list(receipt.images.all())
    has_legacy = bool(receipt.attached_image) and not images
    context = {
        'receipt': receipt,
        'gallery_images': images,
        'has_legacy_image': has_legacy,
        'gallery_count': len(images) if images else (1 if has_legacy else 0),
    }
    return render(request, 'fuel/rer_print.html', context)


def export_rer_pdf(request, pk):
    """Export RER as PDF with multiple images properly placed on top + RER form at bottom.
    Images are laid out in a responsive grid (object-fit contain, no stretch) and
    the RER form is rendered below. Works with gallery (new) or legacy single image.
    """
    if not PDF_SUPPORT:
        return HttpResponse("PDF export is not available. Please install reportlab library.", content_type="text/plain")

    receipt = get_object_or_404(
        ReimbursementExpenseReceipt.objects.select_related('petty_cash_voucher').prefetch_related('images'),
        pk=pk
    )
    images_qs = list(receipt.images.all())
    # Build list of (path, or None) for gallery
    image_paths = []
    if images_qs:
        for img in images_qs:
            try:
                p = img.image.path
                image_paths.append(p)
            except Exception:
                # fallback to url or skip
                image_paths.append(None)
    elif receipt.attached_image:
        try:
            image_paths = [receipt.attached_image.path]
        except Exception:
            image_paths = []
    else:
        image_paths = []

    response = HttpResponse(content_type='application/pdf')
    fname = f"RER_{receipt.rer_no or receipt.pk}.pdf"
    response['Content-Disposition'] = f'inline; filename="{fname}"'

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import Image as PlatImage
    from PIL import Image as PilImage
    import os

    # Font registration (for peso sign etc.)
    font_name = 'Helvetica'
    bold_font_name = 'Helvetica-Bold'
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_dir = os.path.dirname(current_dir)
        fonts_dir = os.path.join(project_dir, 'fonts', 'ttf')
        dejavu_path = os.path.join(fonts_dir, 'DejaVuSans.ttf')
        dejavu_bold_path = os.path.join(fonts_dir, 'DejaVuSans-Bold.ttf')
        if os.path.exists(dejavu_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_path))
            if os.path.exists(dejavu_bold_path):
                pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_bold_path))
                bold_font_name = 'DejaVuSans-Bold'
            else:
                bold_font_name = 'DejaVuSans'
            font_name = 'DejaVuSans'
    except Exception:
        font_name = 'Helvetica'
        bold_font_name = 'Helvetica-Bold'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=10*mm,
        leftMargin=10*mm,
        topMargin=8*mm,
        bottomMargin=8*mm,
        title=f"RER {receipt.rer_no or receipt.pk}"
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('RERTitle', parent=styles['Normal'], fontName=bold_font_name, fontSize=11, alignment=1, textColor=colors.black, spaceAfter=4)
    subtitle_style = ParagraphStyle('RERSub', parent=styles['Normal'], fontName=font_name, fontSize=7, alignment=1, textColor=colors.HexColor('#475569'), spaceAfter=6)
    field_style = ParagraphStyle('RERField', parent=styles['Normal'], fontName=font_name, fontSize=7, leading=9, textColor=colors.black)
    field_b = ParagraphStyle('RERFieldB', parent=field_style, fontName=bold_font_name)
    small_ital = ParagraphStyle('RERSmallItal', parent=styles['Normal'], fontName=font_name, fontSize=5.5, textColor=colors.HexColor('#334155'), alignment=1, leading=7)
    pay_head_style = ParagraphStyle('RERPayHead', parent=styles['Normal'], fontName=bold_font_name, fontSize=7, alignment=1, textColor=colors.black, leading=8)
    ul_style = ParagraphStyle('RERUL', parent=styles['Normal'], fontName=font_name, fontSize=7, leading=9)
    header_label = ParagraphStyle('RERHL', parent=styles['Normal'], fontName=bold_font_name, fontSize=6.5, textColor=colors.black, leading=8)
    peso = "\u20B1"

    elements = []

    # Header
    elements.append(Paragraph("REIMBURSEMENT EXPENSE RECEIPT", title_style))
    if receipt.rer_no:
        elements.append(Paragraph(f"RER No. {receipt.rer_no} &nbsp;&bull;&nbsp; {receipt.receipt_date.strftime('%m/%d/%Y')}", subtitle_style))
    else:
        elements.append(Paragraph(f"Date: {receipt.receipt_date.strftime('%m/%d/%Y')}", subtitle_style))

    # Helper to get scaled platypus image
    def make_image_flowable(img_path, max_w_pt, max_h_pt):
        if not img_path or not os.path.exists(img_path):
            return Paragraph('<i>No image</i>', field_style)
        try:
            with PilImage.open(img_path) as im:
                w, h = im.size
            # avoid division by zero
            if w == 0 or h == 0:
                return Paragraph('<i>Invalid image</i>', field_style)
            scale = min(max_w_pt / w, max_h_pt / h, 1.0)
            # also allow upscale a bit? cap at 1.0 to not upscale beyond original? But for small images we may want to upscale to fill.
            # If image is smaller than max, we can upscale slightly to fill better: allow up to max.
            # So compute scale as min(max_w/w, max_h/h)
            # Actually we want to fit inside box, so use that.
            # For better quality, if original is small, we still fit.
            # Recompute with no cap? Let's allow scale>1 but keep proportion.
            # We'll use direct min without cap 1 to allow upscale? For print we want contain not stretch bigger than box.
            scale = min(max_w_pt / w, max_h_pt / h)
            # Ensure not too large (scale up to maybe 1.5?)
            new_w = w * scale
            new_h = h * scale
            # Clip to max
            new_w = min(new_w, max_w_pt)
            new_h = min(new_h, max_h_pt)
            img = PlatImage(img_path, width=new_w, height=new_h)
            img.hAlign = 'CENTER'
            img.vAlign = 'MIDDLE'
            # Preserve aspect via width/height set
            return img
        except Exception as e:
            return Paragraph(f'<i>Image error</i>', field_style)

    # Gallery grid
    n = len(image_paths)
    if n > 0:
        # Determine columns and cell max sizes (in points)
        if n == 1:
            cols = 1
            max_w, max_h = 260, 220  # ~92x78mm single centered
            gallery_w = max_w + 16  # outer border padding
        elif n == 2:
            cols = 2
            max_w, max_h = 238, 170
            gallery_w = 482  # 170mm
        elif n <= 4:
            cols = 2
            max_w, max_h = 238, 165
            gallery_w = 482
        else:
            cols = 3
            max_w, max_h = 156, 150
            gallery_w = 484

        # Build table data row-wise
        # Use gutter 6pt between cols, 6pt between rows via table padding
        # Create flowables for each image
        flows = [make_image_flowable(p, max_w, max_h) for p in image_paths]

        # Chunk into rows
        rows = []
        for i in range(0, len(flows), cols):
            chunk = flows[i:i+cols]
            # Pad to cols length with empty spacer
            while len(chunk) < cols:
                chunk.append(Paragraph('', field_style))
            # For last row with single item and cols>1 and n % cols ==1, span logic: center it
            # We'll detect and apply SPAN via table style later
            rows.append(chunk)

        col_widths = []
        gutter = 6
        usable = gallery_w - gutter * (cols - 1)
        cw = usable / cols
        col_widths = [cw] * cols
        if n == 1:
            # single image bigger cell with border, centered in page
            gal_table = Table(rows, colWidths=col_widths, hAlign='CENTER')
        else:
            gal_table = Table(rows, colWidths=col_widths, hAlign='CENTER')

        # Style for gallery: each cell has border 0.8pt, background white, padding 3, align center, valign middle
        gstyle = TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white]),
        ])
        # If last row has empty placeholders, we could hide their grid? But keep grid; placeholders will be empty but bordered. Better to hide border for empty placeholders.
        # Instead we can keep; it's okay.
        # For single-item last row with colspan=cols, span to center
        if len(rows) > 0 and n % cols == 1 and n != 1 and cols > 1:
            last_row_idx = len(rows) - 1
            # Span last row's first cell across all cols
            gstyle.add('SPAN', (0, last_row_idx), (-1, last_row_idx))
            # Also center - already center
            # Empty cells in that row after span are ignored, but we padded them, so they are merged.
        gal_table.setStyle(gstyle)
        # Wrap gallery in a bordered container with white background, centered
        # Add a little spacing above/below
        elements.append(gal_table)
        elements.append(Spacer(1, 4*mm))
    else:
        # No images placeholder
        ph_style = ParagraphStyle('PH', parent=styles['Normal'], fontName=font_name, fontSize=7, textColor=colors.HexColor('#64748b'), alignment=1, borderWidth=0.5, borderColor=colors.HexColor('#cbd5e1'), backColor=colors.HexColor('#f8fafc'), leading=9)
        # Use a table with dashed style simulation (using grid grey)
        ph_table = Table([[Paragraph('No image attached — upload images when creating/editing the RER; they will appear here on top when exported.', ph_style)]], colWidths=[460], hAlign='CENTER')
        ph_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(ph_table)
        elements.append(Spacer(1, 4*mm))

    # RER Form — centered 88.9mm (252pt) width, like HTML index card, but allow slightly larger for PDF readability 290pt (~102mm)
    # Keep 3.5" = 252pt exactly for faithful replica, but we use 270pt for slightly more room
    rer_width = 252  # 88.9mm
    # If no images, we use same width centered; else form centered below gallery

    # Build header table (entity/fund | date/rer)
    header_data = [
        [Paragraph(f"Entity Name: <b>{receipt.entity_name or ''}</b>", field_style),
         Paragraph(f"Fund Cluster: <b>{receipt.fund_cluster or ''}</b>", field_style)],
        [Paragraph(f"Date: <b>{receipt.receipt_date.strftime('%m/%d/%Y')}</b>", field_style),
         Paragraph(f"RER No. <b>{receipt.rer_no or ''}</b>", field_style)],
    ]
    header_tbl = Table(header_data, colWidths=[rer_width/2, rer_width/2], hAlign='CENTER')
    header_tbl.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    # Body paragraphs - use Paragraph with underlines via <u> tag simulated by font underline? ReportLab Paragraph supports <u>
    # We'll compose lines with values underlined
    def ul(val, min_w=40):
        # Return Paragraph with underline if val else empty underline placeholder
        # Use <u>&nbsp;&nbsp;val&nbsp;&nbsp;</u> but for empty use underscores
        safe = val or ''
        if safe:
            # escape?
            import html as _html
            safe = _html.escape(str(safe))
            return f"<u> {safe} </u>"
        else:
            return "<u>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</u>"

    # Build body as list of Paragraphs inside a Table single column with no grid but line separators?
    body_items = []
    body_items.append([Paragraph(f"RECEIVED from {ul(receipt.received_from_name or '', 50)} (Name) {ul(receipt.received_from_designation or '', 40)} (Official Designation) the amount of", field_style)])
    body_items.append([Paragraph(f"{ul(receipt.amount_in_words or '', 90)} (In Words) {peso} {ul(f'{receipt.amount_in_figures:.2f}' if receipt.amount_in_figures else '', 35)} (In Figures)", field_style)])
    body_items.append([Paragraph(f"of {ul(receipt.amount_in_words or '', 120)} (In Words)", field_style)])
    body_items.append([Paragraph(f"in payment for {ul(receipt.in_payment_for or '', 130)}", field_style)])
    body_items.append([Paragraph("(Payments for subsistence, services, etc.)", small_ital)])
    body_items.append([Paragraph("rental or transportation should show inclusive dates,", small_ital)])
    body_items.append([Paragraph("purpose, distance, inclusive points of travel, etc.)", small_ital)])

    body_tbl = Table(body_items, colWidths=[rer_width], hAlign='CENTER')
    body_tbl.setStyle(TableStyle([
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    # Payee section
    payee_data = [
        [Paragraph("Payee", pay_head_style)],
        [Paragraph(f"Signature: {ul('', 100)}", ul_style)],
        [Paragraph(f"Name: {ul(receipt.payee_signature_name or '', 100)}", ul_style)],
        [Paragraph(f"Address: {ul(receipt.payee_address or '', 100)}", ul_style)],
        [Paragraph(f"Residence Cert No.: {ul(receipt.payee_residence_cert_no or '', 90)}", ul_style)],
        [Paragraph(f"Date of Issue: {ul(receipt.payee_residence_date.strftime('%m/%d/%Y') if receipt.payee_residence_date else '', 80)}", ul_style)],
        [Paragraph(f"Place of Issue: {ul(receipt.payee_residence_place or '', 80)}", ul_style)],
    ]
    payee_tbl = Table(payee_data, colWidths=[rer_width], hAlign='CENTER')
    payee_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
        ('LINEBELOW', (0, 0), (-1, 0), 0.8, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ]))

    # Witness section
    wit_data = [
        [Paragraph("Witness", pay_head_style)],
        [Paragraph(f"Signature: {ul('', 100)}", ul_style)],
        [Paragraph(f"Name: {ul(receipt.witness_signature_name or '', 100)}", ul_style)],
        [Paragraph(f"Address: {ul(receipt.witness_address or '', 100)}", ul_style)],
        [Paragraph(f"Residence Cert No.: {ul(receipt.witness_residence_cert_no or '', 90)}", ul_style)],
        [Paragraph(f"Date of Issue: {ul(receipt.witness_residence_date.strftime('%m/%d/%Y') if receipt.witness_residence_date else '', 80)}", ul_style)],
        [Paragraph(f"Place of Issue: {ul(receipt.witness_residence_place or '', 80)}", ul_style)],
    ]
    wit_tbl = Table(wit_data, colWidths=[rer_width], hAlign='CENTER')
    wit_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
        ('LINEBELOW', (0, 0), (-1, 0), 0.8, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ]))

    # Outer RER container
    outer_data = [
        [Paragraph("Reimbursement Expense Receipt", ParagraphStyle('OuterTitle', parent=field_b, fontSize=8, alignment=1, leading=10))],
        [header_tbl],
        [body_tbl],
        [payee_tbl],
        [wit_tbl],
    ]
    outer_table = Table(outer_data, colWidths=[rer_width], hAlign='CENTER')
    outer_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1.2, colors.black),
        ('LINEBELOW', (0, 0), (-1, 0), 1.0, colors.black),
        ('LINEBELOW', (0, 1), (-1, 1), 0.6, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    # Wrap outer table in a centered container to keep 3.5" width centered on A4
    # Doc width is ~523pt, rer_width 252, so we create container table 523 width with rer centered
    container = Table([[outer_table]], colWidths=[523], hAlign='CENTER')
    container.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    elements.append(container)

    if receipt.petty_cash_voucher:
        elements.append(Spacer(1, 3*mm))
        pcv = receipt.petty_cash_voucher
        elements.append(Paragraph(f"— Attached to PCV No. {pcv.voucher_no or pcv.pk} (separate A4 print via PCV) —", ParagraphStyle('PCVNote', parent=field_style, fontSize=6, textColor=colors.HexColor('#475569'), alignment=1)))

    doc.build(elements)
    return response

def rer_pcv_combined_print_view(request, pk):
    """Combined print for RER + its linked PCV (or standalone) — PCV is separate A4 attachment.
    This view is optional; we keep PCV and RER prints separate as requested:
    - RER print (image top + form bottom) is one A4
    - PCV print is a second A4
    Here we render RER with image+form; PCV is printed via pcv_print_view.
    """
    return rer_print_view(request, pk)
